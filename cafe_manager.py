import tkinter as tk
from tkinter import ttk, messagebox
import json
import math
import os
import re
import threading
import urllib.request
import urllib.error
from datetime import datetime, date
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── paths ──────────────────────────────────────────────────────────────────────
BASE_DIR       = os.path.dirname(os.path.abspath(__file__))
CONFIG_FILE    = os.path.join(BASE_DIR, "config.json")
DATA_DIR       = os.path.join(BASE_DIR, "data")
CAFE_DATA_FILE = os.path.join(DATA_DIR, "cafe_data.json")
os.makedirs(DATA_DIR, exist_ok=True)

DEFAULT_CONFIG = {
    "rate_0:30":         50,
    "rate_1:00":         90,
    "rate_1:30":         140,
    "rate_2:00":         180,
    "rate_2:30":         230,
    "rate_3:00":         270,
    "rate_4:00":         360,
    "rate_5:00":         450,
    "openrouter_api_key": "",
}

# ── colour palette ─────────────────────────────────────────────────────────────
BG_MAIN    = "#0d1117"
BG_PANEL   = "#161b22"
BG_CARD    = "#21262d"
BG_INPUT   = "#010409"
ACCENT     = "#f85149"
ACCENT_H   = "#ff6b6b"   # accent hover
ACCENT2    = "#8b5cf6"
ACCENT2_H  = "#a78bfa"
BTN_GREEN  = "#238636"
BTN_GREEN_H= "#2ea043"
TEXT_H     = "#e6edf3"
TEXT_MAIN  = "#c9d1d9"
TEXT_DIM   = "#8b949e"
BORDER     = "#30363d"
ROW_ODD    = "#161b22"
ROW_EVEN   = "#0d1117"
ROW_SEL    = "#1f3a5f"

# PC status colours
PC_FREE    = "#1a7f37"
PC_FREE_FG = "#3fb950"
PC_BUSY    = "#8b1a1a"
PC_BUSY_FG = "#f85149"
PC_WARN    = "#7d4a00"
PC_WARN_FG = "#e3b341"
PC_EXP_A   = "#f85149"
PC_EXP_B   = "#e3b341"
PC_OPEN    = "#7c3900"    # deep orange – open session background
PC_OPEN_FG = "#fb923c"    # orange text for open sessions
PC_OFF     = "#0a0a0a"    # near-black – shutdown PC
PC_OFF_FG  = "#3a3a3a"    # dark grey text for shutdown

# Monitor canvas colours (shared between build and glow tick)
MON_BEZEL  = "#1a1f27"    # outer bezel / canvas background
MON_STAND  = "#2a2f38"    # neck + base of monitor stand

# Session-mode colours
BTN_ORANGE   = "#c2410c"
BTN_ORANGE_H = "#ea580c"

# Duration preset buttons
DURATION_PRESETS = [
    ("30 min", "0:30"),
    ("1 hr",   "1:00"),
    ("1:30",   "1:30"),
    ("2 hr",   "2:00"),
    ("2:30",   "2:30"),
    ("3 hr",   "3:00"),
    ("4 hr",   "4:00"),
    ("5 hr",   "5:00"),
]
DURATION_OPTS = [d[1] for d in DURATION_PRESETS]



# ── Colour helpers ─────────────────────────────────────────────────────────────

def _blend_hex(c1: str, c2: str, t: float) -> str:
    """Linearly blend two #rrggbb colours. t=0 → c1, t=1 → c2."""
    t  = max(0.0, min(1.0, t))
    r1, g1, b1 = int(c1[1:3], 16), int(c1[3:5], 16), int(c1[5:7], 16)
    r2, g2, b2 = int(c2[1:3], 16), int(c2[3:5], 16), int(c2[5:7], 16)
    return (f"#{int(r1 + (r2-r1)*t):02x}"
            f"{int(g1 + (g2-g1)*t):02x}"
            f"{int(b1 + (b2-b1)*t):02x}")


# ══════════════════════════════════════════════════════════════════════════════
# CONFIG
# ══════════════════════════════════════════════════════════════════════════════

def load_config():
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE) as f:
                cfg = json.load(f)
            for k, v in DEFAULT_CONFIG.items():
                cfg.setdefault(k, v)
            return cfg
        except Exception:
            pass
    return DEFAULT_CONFIG.copy()


def save_config(cfg):
    with open(CONFIG_FILE, "w") as f:
        json.dump(cfg, f, indent=2)


# ══════════════════════════════════════════════════════════════════════════════
# LOCAL JSON PERSISTENCE  (internal app state — separate from Excel)
# ══════════════════════════════════════════════════════════════════════════════

def load_cafe_data():
    """Return (sessions, expenses, shutdown_pcs, bookings, edit_log) for today."""
    today_str = date.today().isoformat()
    if not os.path.exists(CAFE_DATA_FILE):
        return [], [], set(), [], []
    try:
        with open(CAFE_DATA_FILE) as f:
            data = json.load(f)
    except Exception:
        return [], [], set(), [], []

    if data.get("date") != today_str:
        # Stale data from a previous day — discard it.
        try:
            os.remove(CAFE_DATA_FILE)
        except Exception:
            pass
        return [], [], set(), [], []

    # Only restore pending bookings (started/cancelled don't need to reappear)
    bookings = [b for b in data.get("bookings", [])
                if b.get("status") == "pending"]
    return (data.get("sessions", []),
            data.get("expenses", []),
            set(data.get("shutdown_pcs", [])),
            bookings,
            data.get("edit_log", []))


def save_cafe_data(records, expenses, shutdown_pcs=None, bookings=None,
                   edit_log=None):
    """Persist today's session records, expenses, shutdown state, bookings, edit log."""
    data = {
        "date":         date.today().isoformat(),
        "sessions":     records,
        "expenses":     expenses,
        "shutdown_pcs": list(shutdown_pcs or []),
        "bookings":     bookings or [],
        "edit_log":     edit_log or [],
    }
    try:
        with open(CAFE_DATA_FILE, "w") as f:
            json.dump(data, f, indent=2)
    except Exception:
        pass   # non-critical; Excel is still the primary record


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def excel_path():
    return os.path.join(DATA_DIR, date.today().strftime("%b-%Y") + ".xlsx")


def sheet_name():
    return date.today().strftime("%d-%b")


# Columns: PC No | Name | Time In | Time Out | Amount | Comment  (6 total)
# ID is internal only — kept in cafe_data.json, never written to Excel.
HEADER_COLS = ["PC No", "Name", "Time In", "Time Out", "Amount", "Comment"]
COL_WIDTHS  = [7,       22,     9,         9,          11,        30]

# Dark-red manual-record colour scheme
HEADER_FILL = PatternFill("solid", fgColor="6B1414")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
ODD_FILL    = PatternFill("solid", fgColor="3D1010")
EVEN_FILL   = PatternFill("solid", fgColor="2A0A0A")
TOTAL_FILL  = PatternFill("solid", fgColor="C0392B")
TOTAL_FONT  = Font(bold=True, color="FFFFFF", size=11)
XL_CENTER   = Alignment(horizontal="center", vertical="center")
XL_LEFT     = Alignment(horizontal="left",   vertical="center")
THIN        = Side(style="thin", color="6B1414")
XL_BORDER   = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Expense section colour scheme (dark blue)
EXP_HDR_FILL = PatternFill("solid", fgColor="0e2a47")
EXP_HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
EXP_ODD_FILL = PatternFill("solid", fgColor="0a1f36")
EXP_EVN_FILL = PatternFill("solid", fgColor="071629")
EXP_TOT_FILL = PatternFill("solid", fgColor="1d5b8e")
EXP_TOT_FONT = Font(bold=True, color="FFFFFF", size=11)
EXP_REM_FILL = PatternFill("solid", fgColor="145214")
EXP_THIN     = Side(style="thin", color="1d5b8e")
EXP_BORDER   = Border(left=EXP_THIN, right=EXP_THIN,
                       top=EXP_THIN,  bottom=EXP_THIN)

_NUM_COLS   = len(HEADER_COLS)   # 5
_LAST_COL   = get_column_letter(_NUM_COLS)   # "E"


def _fmt_time_xl(time_str):
    """Convert '09:30 AM' → '9:30' (24-hour, no leading zero, no AM/PM).
    Returns the original string unchanged if it is 'OPEN', '—', or unparseable.
    """
    if not time_str or time_str in ("OPEN", "—", ""):
        return time_str
    try:
        t = datetime.strptime(time_str.strip(), "%I:%M %p")
        return f"{t.hour}:{t.minute:02d}"
    except Exception:
        return time_str


def next_record_id(records):
    """Generate the next YYYYMMDD-XXX id, based on existing records for today."""
    prefix = date.today().strftime("%Y%m%d") + "-"
    max_seq = 0
    for r in records:
        rid = r.get("id", "")
        if rid.startswith(prefix):
            try:
                max_seq = max(max_seq, int(rid[len(prefix):]))
            except ValueError:
                pass
    return f"{prefix}{max_seq + 1:03d}"


def _ensure_sheet(wb, sname, today_label):
    if sname in wb.sheetnames:
        return wb[sname]
    ws = wb.create_sheet(sname)
    ws.merge_cells(f"A1:{_LAST_COL}1")
    ws["A1"] = f"Internet Cafe — {today_label}"
    ws["A1"].font      = Font(bold=True, color="FFFFFF", size=13)
    ws["A1"].fill      = PatternFill("solid", fgColor="8B1A1A")
    ws["A1"].alignment = XL_CENTER
    ws.row_dimensions[1].height = 22
    for col, (h, w) in enumerate(zip(HEADER_COLS, COL_WIDTHS), start=1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = HEADER_FONT; cell.fill = HEADER_FILL
        cell.alignment = XL_CENTER; cell.border = XL_BORDER
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[2].height = 18
    return ws


def _rewrite_sheet(ws, records, expenses):
    # Unmerge everything first — stale merged ranges (e.g. from the previous
    # total row) cause openpyxl to silently skip writes to slave cells (B, C, D),
    # which is why those columns go blank on the 2nd+ save.
    for rng in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(rng))
    if ws.max_row >= 3:
        ws.delete_rows(3, ws.max_row - 2)
    for i, rec in enumerate(records):
        row  = i + 3
        fill = ODD_FILL if i % 2 == 0 else EVEN_FILL
        is_open = rec.get("session_type") == "open" and rec.get("time_out") == "OPEN"
        # Exactly 5 values — one per column, no ID in Excel
        values = [
            rec["pc"],
            rec["name"],
            _fmt_time_xl(rec["time_in"]),
            "OPEN" if is_open else _fmt_time_xl(rec["time_out"]),
            "OPEN" if is_open else rec["final"],
            rec.get("comment", ""),
        ]
        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = fill; cell.border = XL_BORDER
            cell.font = Font(color="FFFFFF", size=10)
            cell.alignment = XL_LEFT if col == 2 else XL_CENTER
        ws.row_dimensions[row].height = 16

    # Total row — only count closed (non-open) sessions
    total_row   = len(records) + 3
    closed      = [r for r in records
                   if not (r.get("session_type") == "open"
                           and r.get("time_out") == "OPEN")]
    total_final = sum(r["final"] for r in closed)
    ws.merge_cells(f"A{total_row}:D{total_row}")   # merge cols 1–4
    c  = ws.cell(row=total_row, column=1, value="TOTAL EARNINGS")
    c.font = TOTAL_FONT; c.fill = TOTAL_FILL
    c.alignment = XL_CENTER; c.border = XL_BORDER
    c2 = ws.cell(row=total_row, column=5, value=total_final)   # col 5 = Amount
    c2.font = TOTAL_FONT; c2.fill = TOTAL_FILL
    c2.alignment = XL_CENTER; c2.border = XL_BORDER
    ws.row_dimensions[total_row].height = 18

    # ── Expenses section (written below the sessions block) ───────────────────
    if expenses:
        exp_hdr_row = total_row + 1
        ws.merge_cells(f"A{exp_hdr_row}:{_LAST_COL}{exp_hdr_row}")
        ch = ws.cell(row=exp_hdr_row, column=1, value="EXPENSES")
        ch.font = EXP_HDR_FONT; ch.fill = EXP_HDR_FILL
        ch.alignment = XL_CENTER; ch.border = EXP_BORDER
        ws.row_dimensions[exp_hdr_row].height = 16

        for j, exp in enumerate(expenses):
            er = exp_hdr_row + 1 + j
            fill_e = EXP_ODD_FILL if j % 2 == 0 else EXP_EVN_FILL
            ws.merge_cells(f"A{er}:D{er}")
            cn = ws.cell(row=er, column=1, value=exp["name"])
            cn.font = Font(color="FFFFFF", size=10)
            cn.fill = fill_e; cn.alignment = XL_LEFT; cn.border = EXP_BORDER
            ca = ws.cell(row=er, column=5, value=exp["amount"])
            ca.font = Font(color="FFFFFF", size=10)
            ca.fill = fill_e; ca.alignment = XL_CENTER; ca.border = EXP_BORDER
            ws.row_dimensions[er].height = 16

        total_exp = sum(e["amount"] for e in expenses)
        texp_row  = exp_hdr_row + 1 + len(expenses)
        ws.merge_cells(f"A{texp_row}:D{texp_row}")
        ct = ws.cell(row=texp_row, column=1, value="TOTAL EXPENSES")
        ct.font = EXP_TOT_FONT; ct.fill = EXP_TOT_FILL
        ct.alignment = XL_CENTER; ct.border = EXP_BORDER
        ct2 = ws.cell(row=texp_row, column=5, value=total_exp)
        ct2.font = EXP_TOT_FONT; ct2.fill = EXP_TOT_FILL
        ct2.alignment = XL_CENTER; ct2.border = EXP_BORDER
        ws.row_dimensions[texp_row].height = 18

        rem_row = texp_row + 1
        ws.merge_cells(f"A{rem_row}:D{rem_row}")
        cr = ws.cell(row=rem_row, column=1, value="REMAINING")
        cr.font = EXP_TOT_FONT; cr.fill = EXP_REM_FILL
        cr.alignment = XL_CENTER; cr.border = EXP_BORDER
        cr2 = ws.cell(row=rem_row, column=5,
                       value=max(0, total_final - total_exp))
        cr2.font = EXP_TOT_FONT; cr2.fill = EXP_REM_FILL
        cr2.alignment = XL_CENTER; cr2.border = EXP_BORDER
        ws.row_dimensions[rem_row].height = 18


def save_to_excel(records, expenses):
    path        = excel_path()
    sname       = sheet_name()
    today_label = date.today().strftime("%B %d, %Y")
    wb = openpyxl.load_workbook(path) if os.path.exists(path) else openpyxl.Workbook()
    if not os.path.exists(path) and "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    ws = _ensure_sheet(wb, sname, today_label)
    _rewrite_sheet(ws, records, expenses)
    idx = wb.sheetnames.index(sname)
    wb.move_sheet(sname, offset=-idx)
    wb.save(path)


def migrate_excel_format():
    """Delete stale Excel and JSON data if the column layout is outdated.

    Triggers when the sheet header contains any of the old or transitional
    column names (Duration, Discount, Final Amount, ID).  Both the Excel file
    and cafe_data.json are wiped so the app starts completely fresh.
    """
    _STALE_HEADERS = {"Duration", "Discount", "Final Amount", "ID"}
    path = excel_path()
    needs_wipe = False

    if os.path.exists(path):
        try:
            wb    = openpyxl.load_workbook(path, read_only=True)
            sname = sheet_name()
            if sname in wb.sheetnames:
                ws      = wb[sname]
                headers = {ws.cell(row=2, column=c).value
                           for c in range(1, 10)}
                if headers & _STALE_HEADERS:
                    needs_wipe = True
            wb.close()
        except Exception:
            needs_wipe = True   # unreadable file — wipe it too

    if needs_wipe:
        for f in (path, CAFE_DATA_FILE):
            try:
                os.remove(f)
            except Exception:
                pass


# ══════════════════════════════════════════════════════════════════════════════
# CALCULATION HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def calc_amount(minutes, cfg):
    """Fallback amount calculator for custom durations not matching a preset."""
    if minutes <= 0:
        return 0
    rate_hour  = cfg.get("rate_1:00",  90)
    rate_30min = cfg.get("rate_0:30",  50)
    hours    = minutes // 60
    leftover = minutes % 60
    amount   = hours * rate_hour
    if leftover > 0:
        amount += ((leftover + 29) // 30) * rate_30min
    return amount


def fmt_duration(minutes):
    if minutes <= 0: return "0 min"
    h, m = minutes // 60, minutes % 60
    if h and m: return f"{h}h {m}m"
    return f"{h}h" if h else f"{m}m"


def parse_dur_input(s):
    try:
        parts = s.strip().split(":")
        return int(parts[0]) * 60 + (int(parts[1]) if len(parts) > 1 else 0)
    except Exception:
        return 0


def minutes_to_dur_input(minutes):
    return f"{minutes // 60}:{minutes % 60:02d}"


def calc_timeout_str(time_in_minutes, duration_minutes):
    total = (time_in_minutes + duration_minutes) % (24 * 60)
    h24   = total // 60
    m     = total % 60
    h12   = h24 % 12 or 12
    return f"{h12:02d}:{m:02d} {'AM' if h24 < 12 else 'PM'}"


def parse_session_time(time_str):
    try:
        t = datetime.strptime(time_str.strip(), "%I:%M %p")
        return datetime.combine(date.today(), t.time())
    except Exception:
        return None


# ══════════════════════════════════════════════════════════════════════════════
# UI HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def hover_bind(widget, normal_bg, hover_bg, also=()):
    """Attach hover colour-swap to widget and optional child widgets."""
    def on_enter(_):
        widget.config(bg=hover_bg)
        for w in also: w.config(bg=hover_bg)
    def on_leave(_):
        widget.config(bg=normal_bg)
        for w in also: w.config(bg=normal_bg)
    widget.bind("<Enter>", on_enter)
    widget.bind("<Leave>", on_leave)


def section_label(parent, text):
    tk.Label(parent, text=text, bg=BG_PANEL, fg=TEXT_DIM,
             font=("Segoe UI", 7, "bold")).pack(anchor="w", padx=12, pady=(8, 2))
    tk.Frame(parent, bg=BORDER, height=1).pack(fill="x", padx=12, pady=(0, 4))


# ══════════════════════════════════════════════════════════════════════════════
# SETTINGS DIALOG
# ══════════════════════════════════════════════════════════════════════════════

class SettingsDialog(tk.Toplevel):
    _RATE_FIELDS = [
        ("30 min", "rate_0:30"),
        ("1 hr",   "rate_1:00"),
        ("1:30",   "rate_1:30"),
        ("2 hr",   "rate_2:00"),
        ("2:30",   "rate_2:30"),
        ("3 hr",   "rate_3:00"),
        ("4 hr",   "rate_4:00"),
        ("5 hr",   "rate_5:00"),
    ]

    def __init__(self, parent, cfg, on_save):
        super().__init__(parent)
        self.title("Settings")
        self.configure(bg=BG_MAIN)
        self.resizable(False, False)
        self.grab_set()
        self._cfg     = cfg
        self._on_save = on_save
        self._vars    = {}

        tk.Label(self, text="⚙  Duration Rates", bg=BG_MAIN, fg=TEXT_H,
                 font=("Segoe UI", 12, "bold")).grid(
            row=0, column=0, columnspan=2, padx=20, pady=(16, 2), sticky="w")
        tk.Label(self, text="Price charged when a duration button is selected",
                 bg=BG_MAIN, fg=TEXT_DIM, font=("Segoe UI", 8)).grid(
            row=1, column=0, columnspan=2, padx=20, pady=(0, 10), sticky="w")

        for row, (label, key) in enumerate(self._RATE_FIELDS, start=2):
            tk.Label(self, text=f"{label}  (₱):", bg=BG_MAIN, fg=TEXT_MAIN,
                     font=("Segoe UI", 10), width=12, anchor="w").grid(
                row=row, column=0, sticky="w", padx=20, pady=5)
            v = tk.StringVar(value=str(int(cfg.get(key, DEFAULT_CONFIG.get(key, 0)))))
            self._vars[key] = v
            tk.Entry(self, textvariable=v, bg=BG_CARD, fg=TEXT_H,
                     insertbackground=TEXT_H, relief="flat", width=10,
                     font=("Consolas", 11), highlightthickness=1,
                     highlightbackground=BORDER, highlightcolor=ACCENT).grid(
                row=row, column=1, padx=20, pady=5, sticky="w")

        # OpenRouter API key field
        api_row = len(self._RATE_FIELDS) + 2
        tk.Frame(self, bg=BORDER, height=1).grid(
            row=api_row, column=0, columnspan=2, sticky="ew", padx=20, pady=(4, 8))
        tk.Label(self, text="🤖  OpenRouter API Key", bg=BG_MAIN, fg=TEXT_H,
                 font=("Segoe UI", 10, "bold")).grid(
            row=api_row + 1, column=0, columnspan=2, padx=20, pady=(0, 4), sticky="w")
        tk.Label(self, text="openrouter.ai → free account → API keys",
                 bg=BG_MAIN, fg=TEXT_DIM, font=("Segoe UI", 8)).grid(
            row=api_row + 2, column=0, columnspan=2, padx=20, pady=(0, 6), sticky="w")
        self._api_key_var = tk.StringVar(value=cfg.get("openrouter_api_key", ""))
        tk.Entry(self, textvariable=self._api_key_var, bg=BG_CARD, fg=TEXT_H,
                 insertbackground=TEXT_H, relief="flat", width=32,
                 font=("Consolas", 9), highlightthickness=1,
                 highlightbackground=BORDER, highlightcolor=ACCENT2).grid(
            row=api_row + 3, column=0, columnspan=2, padx=20, pady=(0, 4), sticky="ew")

        save_row = api_row + 4
        tk.Button(self, text="  Save Changes  ", command=self._save,
                  bg=BTN_GREEN, fg="white", activebackground=BTN_GREEN_H,
                  relief="flat", font=("Segoe UI", 10, "bold"),
                  padx=20, pady=8, cursor="hand2").grid(
            row=save_row, column=0, columnspan=2, pady=16)
        self.geometry("+%d+%d" % (parent.winfo_rootx()+100, parent.winfo_rooty()+100))

    def _save(self):
        new_vals = {}
        for key, var in self._vars.items():
            try:
                val = float(var.get())
                if val <= 0:
                    raise ValueError
                new_vals[key] = val
            except ValueError:
                messagebox.showerror("Error", "Enter valid positive numbers for all rates.",
                                     parent=self)
                return
        new_vals["openrouter_api_key"] = self._api_key_var.get().strip()
        self._cfg.update(new_vals)
        save_config(self._cfg)
        self._on_save()
        self.destroy()


# ══════════════════════════════════════════════════════════════════════════════
# TIME PICKER
# ══════════════════════════════════════════════════════════════════════════════

class TimePicker(tk.Frame):
    def __init__(self, parent, **kw):
        super().__init__(parent, bg=BG_PANEL, **kw)
        self._hour = tk.StringVar(value="12")
        self._min  = tk.StringVar(value="00")
        self._ampm = tk.StringVar(value="PM")

        spin_kw = dict(bg=BG_CARD, fg=TEXT_H, insertbackground=TEXT_H,
                       relief="flat", width=3, font=("Consolas", 13, "bold"),
                       buttonbackground=BG_CARD, highlightthickness=1,
                       highlightbackground=BORDER, highlightcolor=ACCENT)

        # Validation: allow only digits, max 2 characters
        vcmd_num2 = (self.register(lambda P: P.isdigit() and len(P) <= 2 or P == ""), "P")

        self._sh = tk.Spinbox(self, from_=1, to=12, textvariable=self._hour,
                              format="%02.0f", wrap=True,
                              validate="key", validatecommand=vcmd_num2, **spin_kw)
        self._sh.pack(side="left")
        self._sh.bind("<FocusOut>", self._clamp_hour)
        self._sh.bind("<Return>",   self._clamp_hour)

        tk.Label(self, text=":", bg=BG_PANEL, fg=ACCENT,
                 font=("Consolas", 16, "bold")).pack(side="left", padx=1)

        self._sm = tk.Spinbox(self, from_=0, to=55, textvariable=self._min,
                              format="%02.0f", increment=5, wrap=True,
                              validate="key", validatecommand=vcmd_num2, **spin_kw)
        self._sm.pack(side="left")
        self._sm.bind("<FocusOut>", self._clamp_min)
        self._sm.bind("<Return>",   self._clamp_min)

        tk.Label(self, text=" ", bg=BG_PANEL).pack(side="left", padx=2)

        self._sa = ttk.Combobox(self, textvariable=self._ampm,
                                values=["AM", "PM"], width=3, state="readonly",
                                font=("Segoe UI", 10, "bold"))
        self._sa.pack(side="left")

    def _clamp_hour(self, event=None):
        try:
            v = int(self._hour.get())
        except ValueError:
            v = 0
        if v < 1:
            v = 12       # empty or 0 → default to 12
        elif v > 12:
            v = 12       # over-range → cap at 12
        self._hour.set(f"{v:02d}")

    def _clamp_min(self, event=None):
        try:
            v = int(self._min.get())
        except ValueError:
            v = 0        # empty → 00
        if v > 55:
            v = 55       # over-range → cap at 55
        # snap down to nearest multiple of 5
        v = (v // 5) * 5
        self._min.set(f"{v:02d}")

    def set_now(self):
        now = datetime.now()
        h   = now.hour
        m   = now.minute
        # Ceiling to next 5-minute mark (never round down)
        m = math.ceil(m / 5) * 5
        if m >= 60:
            m  = 0
            h += 1
            if h >= 24:
                h = 0
        h12 = h % 12 or 12
        self._hour.set(f"{h12:02d}")
        self._min.set(f"{m:02d}")
        self._ampm.set("AM" if h < 12 else "PM")

    def get_time(self):
        try:
            h    = int(self._hour.get())
            m    = int(self._min.get())
            ampm = self._ampm.get()
        except ValueError:
            return "??:??", 0
        h24 = h % 12 + (12 if ampm == "PM" else 0)
        return f"{h:02d}:{m:02d} {ampm}", h24 * 60 + m

    def get_str(self):     return self.get_time()[0]
    def get_minutes(self): return self.get_time()[1]



# ══════════════════════════════════════════════════════════════════════════════
# MAIN APP
# ══════════════════════════════════════════════════════════════════════════════

class CafeApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Cafe Manager")
        self.configure(bg=BG_MAIN)
        self.resizable(True, True)
        self.minsize(1060, 820)

        self._cfg            = load_config()
        (self._records, self._expenses, self._shutdown_pcs,
         self._bookings, self._edit_log) = load_cafe_data()
        migrate_excel_format()                    # one-time column-format upgrade
        self._edit_idx        = None
        self._editing_open    = False    # True when editing an existing OPEN record
        self._amount_manual   = False
        self._flash_state     = False
        self._pc_boxes        = {}
        self._form_pc_btns    = {}
        self._dur_btns        = {}
        self._selected_dur    = "1:00"
        self._expire_timers   = {}       # {pc_num: datetime when rem first hit 0}
        self._glow_tick       = 0.0      # phase counter for glow animation
        self._booking_alerted  = set()   # booking IDs already alerted this session
        self._session_alerted  = set()   # record IDs alerted for 5-min warning
        self._notifications    = []      # [{text, color, key}] active notifications
        self._notif_index      = 0       # current rotation index
        self._ai_panel_open   = False
        self._ai_typing       = False
        self._chat_history    = []       # [{role, content, timestamp}]
        self._pending_record  = None     # AI-suggested record awaiting confirmation

        self._build_ui()
        self._refresh_table()
        self._refresh_expenses()
        self._update_summary()
        self.after(0,  self._tick_clock)
        self.after(40, self._tick_glow)

    # ── clock ────────────────────────────────────────────────────────────────
    def _tick_clock(self):
        now = datetime.now()
        self._clock_lbl.config(text=now.strftime("%I:%M:%S %p"))
        self._date_lbl.config(text=now.strftime("%A, %B %d %Y"))
        self._flash_state = not self._flash_state
        self._update_pc_grid()
        self._check_booking_alerts()
        self._check_session_alerts()
        self._sync_notifications()
        self.after(1000, self._tick_clock)

    def _check_session_alerts(self):
        """Add bar notification when a fixed session has 5 min or less remaining."""
        now = datetime.now()
        for rec in self._records:
            if rec.get("session_type") in ("closed", "open"):
                continue
            rid = rec.get("id", "")
            if not rid or rid in self._session_alerted:
                continue
            t_in  = parse_session_time(rec.get("time_in",  ""))
            t_out = parse_session_time(rec.get("time_out", ""))
            if not t_in or not t_out or t_in > now:
                continue
            rem = int((t_out - now).total_seconds())
            if 0 < rem <= 300:
                self._session_alerted.add(rid)
                self._push_notification(
                    text=f"PC {rec['pc']} - {rec['name']} ka time khatam hone wala hai!",
                    color="#fb923c",   # orange
                    key=f"warn_{rid}",
                )
            elif rem <= 0 and rid in self._session_alerted:
                # Upgrade existing warn notification to expired (red)
                self._push_notification(
                    text=f"PC {rec['pc']} - {rec['name']} ka time khatam!",
                    color=ACCENT,      # red
                    key=f"exp_{rid}",
                )

    def _check_booking_alerts(self):
        now = datetime.now()
        now_total = now.hour * 60 + now.minute
        for b in self._bookings:
            if b.get("status") != "pending":
                continue
            bid = b.get("id", "")
            if bid in self._booking_alerted:
                continue
            exp = b.get("exp_time_minutes", -1)
            if 0 <= (now_total - exp) < 5:
                self._booking_alerted.add(bid)
                self._push_notification(
                    text=f"Booking: {b['name']} PC {b['pc']} pe aa gaya!",
                    color="#e3b341",   # yellow
                    key=f"bk_{bid}",
                )

    # ── glow animation loop (40 ms — independent of 1 s clock) ──────────────
    def _tick_glow(self):
        self._glow_tick += 0.08   # phase step: full cycle ≈ 2.5 s at 40 ms

        # (state, min_intensity, max_intensity, speed_multiplier)
        _GLOW = {
            "free":    (0.06, 0.12, 0.0),   # static dim green halo
            "busy":    (0.18, 0.62, 1.0),   # slow breathing red
            "open":    (0.18, 0.55, 1.3),   # medium orange pulse
            "warn":    (0.28, 0.78, 2.2),   # fast yellow pulse
            "expired": (0.35, 0.90, 3.5),   # rapid red flash
        }
        # Layer factors: [outer rect, mid rect, inner rect]
        # inner layer is brightest, outer is most diffuse
        _LAYER = [0.18, 0.45, 0.85]

        for i in range(1, 16):
            info  = self._pc_boxes[i]
            state = info.get("glow_state", "free")
            color = info.get("glow_color", PC_FREE_FG)
            mn, mx, spd = _GLOW.get(state, _GLOW["free"])

            if spd > 0:
                # sin oscillates −1..+1 → map to 0..1
                t         = (math.sin(self._glow_tick * spd) + 1.0) * 0.5
                intensity = mn + (mx - mn) * t
            else:
                intensity = mn   # static

            cv = info["canvas"]
            for gid, layer_factor in zip(info["glow_ids"], _LAYER):
                blended = _blend_hex(MON_BEZEL, color, intensity * layer_factor)
                cv.itemconfig(gid, fill=blended)

        self.after(40, self._tick_glow)

    # ── master layout ────────────────────────────────────────────────────────
    def _build_ui(self):
        self.columnconfigure(0, weight=1)
        self.columnconfigure(1, weight=0)   # AI panel — fixed width, hidden by default
        self.rowconfigure(2, weight=1)

        self._build_header()
        self._build_pc_grid()

        main = tk.Frame(self, bg=BG_MAIN)
        main.grid(row=2, column=0, sticky="nsew")
        main.columnconfigure(1, weight=1)
        main.rowconfigure(0, weight=1)

        self._build_form(main)
        self._build_table(main)
        self._build_summary()
        self._build_ai_panel()          # spans all rows in column=1

    # ── HEADER ───────────────────────────────────────────────────────────────
    def _build_header(self):
        hdr = tk.Frame(self, bg=BG_CARD, height=64)
        hdr.grid(row=0, column=0, sticky="ew")
        hdr.columnconfigure(1, weight=1)
        hdr.grid_propagate(False)

        # logo + title
        logo_frame = tk.Frame(hdr, bg=BG_CARD)
        logo_frame.grid(row=0, column=0, padx=(16, 0), pady=10, sticky="w")

        logo_bg = tk.Frame(logo_frame, bg=ACCENT, width=42, height=42)
        logo_bg.pack(side="left")
        logo_bg.pack_propagate(False)
        tk.Label(logo_bg, text="☕", bg=ACCENT, fg="white",
                 font=("Segoe UI", 20)).place(relx=.5, rely=.5, anchor="center")

        title_f = tk.Frame(logo_frame, bg=BG_CARD)
        title_f.pack(side="left", padx=(10, 0))
        tk.Label(title_f, text="CAFE MANAGER", bg=BG_CARD, fg=TEXT_H,
                 font=("Segoe UI", 14, "bold")).pack(anchor="w")
        tk.Label(title_f, text="Internet Cafe Session Tracker", bg=BG_CARD,
                 fg=TEXT_DIM, font=("Segoe UI", 8)).pack(anchor="w")

        # clock area
        clock_f = tk.Frame(hdr, bg=BG_CARD)
        clock_f.grid(row=0, column=1, sticky="e", padx=16)

        self._clock_lbl = tk.Label(clock_f, text="", bg=BG_CARD, fg=TEXT_H,
                                   font=("Consolas", 20, "bold"))
        self._clock_lbl.pack(anchor="e")
        self._date_lbl = tk.Label(clock_f, text="", bg=BG_CARD, fg=TEXT_DIM,
                                  font=("Segoe UI", 8))
        self._date_lbl.pack(anchor="e")

        # AI Assistant toggle button
        self._ai_toggle_btn = tk.Button(
            hdr, text="🤖  AI Assistant",
            command=self._toggle_ai_panel,
            bg=ACCENT2, fg="white",
            activebackground=ACCENT2_H, activeforeground="white",
            relief="flat", font=("Segoe UI", 9, "bold"),
            padx=12, pady=6, cursor="hand2")
        self._ai_toggle_btn.grid(row=0, column=2, padx=(0, 8))
        hover_bind(self._ai_toggle_btn, ACCENT2, ACCENT2_H)

        # settings button
        settings_btn = tk.Button(hdr, text="⚙  Settings",
                                 command=self._open_settings,
                                 bg=BG_MAIN, fg=TEXT_MAIN,
                                 activebackground=ACCENT2,
                                 activeforeground="white",
                                 relief="flat", font=("Segoe UI", 9),
                                 padx=12, pady=6, cursor="hand2",
                                 highlightthickness=1,
                                 highlightbackground=BORDER)
        settings_btn.grid(row=0, column=3, padx=(0, 16))
        hover_bind(settings_btn, BG_MAIN, ACCENT2)

    # ── PC STATUS GRID ───────────────────────────────────────────────────────
    def _build_pc_grid(self):
        outer = tk.Frame(self, bg=BG_PANEL)
        outer.grid(row=1, column=0, sticky="ew", padx=12, pady=(6, 0))
        outer.columnconfigure(0, weight=1)

        # header row
        hrow = tk.Frame(outer, bg=BG_PANEL)
        hrow.grid(row=0, column=0, sticky="ew", padx=8, pady=(6, 4))
        hrow.columnconfigure(1, weight=1)

        tk.Label(hrow, text="PC STATUS", bg=BG_PANEL, fg=TEXT_DIM,
                 font=("Segoe UI", 8, "bold")).grid(row=0, column=0, sticky="w")

        self._pc_onoff_lbl = tk.Label(hrow, text="ON: 15  |  OFF: 0",
                                      bg=BG_PANEL, fg=TEXT_DIM,
                                      font=("Segoe UI", 8, "bold"))
        self._pc_onoff_lbl.grid(row=0, column=1, sticky="w", padx=(16, 0))

        legend = tk.Frame(hrow, bg=BG_PANEL)
        legend.grid(row=0, column=2, sticky="e")
        for lcol, ltext in (
            (PC_FREE_FG, "Available"),
            (PC_BUSY_FG, "Occupied"),
            (PC_WARN_FG, "< 5 min"),
            (PC_EXP_A,   "Expired"),
        ):
            dot = tk.Frame(legend, bg=lcol, width=8, height=8)
            dot.pack(side="left", padx=(8, 2))
            tk.Label(legend, text=ltext, bg=BG_PANEL, fg=TEXT_DIM,
                     font=("Segoe UI", 7)).pack(side="left", padx=(0, 4))

        # ── Monitor canvas geometry constants ─────────────────────────────
        CV_W, CV_H  = 118, 82          # total canvas size
        SX0, SY0    = 4,  4            # screen top-left
        SX1, SY1    = CV_W - 4, 60    # screen bottom-right  (110 × 56 px)
        CX          = CV_W // 2        # horizontal centre (59)
        # Stand neck
        NX0, NY0    = CX - 6,  SY1
        NX1, NY1    = CX + 6,  SY1 + 10
        # Stand base
        BX0, BY0    = CX - 24, NY1
        BX1, BY1    = CX + 24, CV_H - 2

        # ── Grid: row 0 = PC 1–8, row 1 = PC 9–15 ────────────────────────
        grid_f = tk.Frame(outer, bg=BG_PANEL)
        grid_f.grid(row=1, column=0, sticky="ew", padx=8, pady=(0, 8))
        for c in range(8):
            grid_f.columnconfigure(c, weight=1)

        for i in range(1, 16):
            gr = 0 if i <= 8 else 1
            gc = (i - 1) if i <= 8 else (i - 9)

            cv = tk.Canvas(grid_f, width=CV_W, height=CV_H,
                           bg=MON_BEZEL, highlightthickness=0, cursor="hand2")
            cv.grid(row=gr, column=gc, padx=3, pady=(3, 4), sticky="ew")

            # ── Glow layers (behind the screen rect) ─────────────────────
            # Three concentric rects, expanding 3/2/1 px outside the screen.
            # Colours are updated every 40 ms by _tick_glow().
            glow_ids = []
            for expand in (3, 2, 1):   # outer → inner
                gid = cv.create_rectangle(SX0 - expand, SY0 - expand,
                                          SX1 + expand, SY1 + expand,
                                          fill=MON_BEZEL, outline="", width=0)
                glow_ids.append(gid)

            # ── Screen rect (drawn on top of glow layers) ─────────────────
            scr = cv.create_rectangle(SX0, SY0, SX1, SY1,
                                      fill=PC_FREE, outline=MON_BEZEL, width=2)
            # Stand neck
            cv.create_rectangle(NX0, NY0, NX1, NY1,
                                 fill=MON_STAND, outline="", width=0)
            # Stand base
            cv.create_rectangle(BX0, BY0, BX1, BY1,
                                 fill=MON_STAND, outline="", width=0)

            # ── Text inside the screen ────────────────────────────────────
            t_num   = cv.create_text(CX, 19, text=f"PC {i}",
                                     fill="white",
                                     font=("Segoe UI", 15, "bold"))
            t_name  = cv.create_text(CX, 37, text="",
                                     fill=PC_FREE_FG,
                                     font=("Segoe UI", 9))
            t_timer = cv.create_text(CX, 51, text="Available",
                                     fill=PC_FREE_FG,
                                     font=("Consolas", 9, "bold"))
            # Small "BOOKED" badge — top-right corner of screen, hidden by default
            t_booked = cv.create_text(SX1 - 3, SY0 + 3, text="",
                                      fill=ACCENT2_H,
                                      font=("Segoe UI", 6, "bold"),
                                      anchor="ne")

            cv.bind("<Button-1>", lambda e, n=i: self._on_pc_click(n))
            cv.bind("<Button-3>", lambda e, n=i: self._on_pc_right_click(e, n))

            self._pc_boxes[i] = {
                "canvas":      cv,
                "screen":      scr,
                "glow_ids":    glow_ids,   # [outer, mid, inner]
                "glow_state":  "free",     # set by _update_pc_grid
                "glow_color":  PC_FREE_FG, # primary colour for glow
                "text_num":    t_num,
                "text_name":   t_name,
                "text_timer":  t_timer,
                "text_booked": t_booked,
            }

    def _on_pc_click(self, pc_num):
        if pc_num in self._shutdown_pcs:
            return
        rec, rem = self._get_session_state(pc_num)
        if rec is not None and rem is None:
            # open session → load edit form to close it
            idx = next(i for i, r in enumerate(self._records) if r is rec)
            self._load_open_for_edit(idx, rec)
            return
        self._v_pc.set(str(pc_num))
        self._highlight_form_pc(pc_num)

    def _on_pc_right_click(self, event, pc_num):
        rec, rem = self._get_session_state(pc_num)
        is_occupied = rec is not None
        is_shutdown = pc_num in self._shutdown_pcs
        if is_shutdown:
            menu = tk.Menu(self, tearoff=0)
            menu.configure(bg=BG_CARD, fg=TEXT_H, activebackground=ACCENT2,
                           activeforeground="white", relief="flat", bd=0)
            menu.add_command(label=f"Turn On PC {pc_num}",
                             command=lambda: self._turn_on_pc(pc_num))
            menu.tk_popup(event.x_root, event.y_root)
        elif not is_occupied:
            menu = tk.Menu(self, tearoff=0)
            menu.configure(bg=BG_CARD, fg=TEXT_H, activebackground=ACCENT,
                           activeforeground="white", relief="flat", bd=0)
            menu.add_command(label=f"Shutdown PC {pc_num}",
                             command=lambda: self._shutdown_pc(pc_num))
            menu.tk_popup(event.x_root, event.y_root)

    def _shutdown_pc(self, pc_num):
        self._shutdown_pcs.add(pc_num)
        self._persist()
        self._update_pc_grid()

    def _turn_on_pc(self, pc_num):
        self._shutdown_pcs.discard(pc_num)
        self._persist()
        self._update_pc_grid()

    def _update_add_btn_state(self):
        if self._editing_open or self._edit_idx is not None:
            self._btn_add.config(state="normal")
            return
        try:
            pc = int(self._v_pc.get())
        except (ValueError, AttributeError):
            return
        if pc in self._shutdown_pcs:
            self._btn_add.config(state="disabled", text="⛔  PC IS SHUTDOWN",
                                 bg=BG_CARD, fg=TEXT_DIM)
        else:
            self._btn_add.config(state="normal")

    def _get_session_state(self, pc_num):
        """Return (rec, remaining_seconds) for the active session on pc_num.
        remaining_seconds is None for open sessions, negative for expired,
        0 if no session found."""
        now  = datetime.now()
        best = None
        for rec in self._records:
            if str(rec["pc"]) != str(pc_num): continue
            if rec.get("session_type") == "closed": continue  # already freed
            # open session still running
            if rec.get("session_type") == "open" and rec.get("time_out") == "OPEN":
                t_in = parse_session_time(rec["time_in"])
                if t_in and t_in <= now:
                    return (rec, None)   # None = open/running
            else:
                t_in  = parse_session_time(rec["time_in"])
                t_out = parse_session_time(rec["time_out"])
                if t_in and t_out and t_in <= now:
                    rem = int((t_out - now).total_seconds())
                    if best is None or rem > best[1]:
                        best = (rec, rem)
        return best if best else (None, 0)

    def _get_active_session(self, pc_num):
        rec, rem = self._get_session_state(pc_num)
        if rec is None: return None
        if rem is None: return rec    # open session is always active
        return rec if rem > 0 else None

    def _update_pc_grid(self):
        now          = datetime.now()
        need_persist = False

        for i in range(1, 16):
            info = self._pc_boxes[i]

            if i in self._shutdown_pcs:
                bg, fg_sub          = PC_OFF, PC_OFF_FG
                name_txt, timer_txt = "", "OFF"
                glow_state, glow_color = "free", PC_OFF_FG
                self._expire_timers.pop(i, None)
            else:
                rec, rem = self._get_session_state(i)

                if rec is None:
                    bg, fg_sub              = PC_FREE, PC_FREE_FG
                    name_txt, timer_txt     = "", "Available"
                    glow_state, glow_color  = "free", PC_FREE_FG
                    self._expire_timers.pop(i, None)

                elif rem is None:                              # open session
                    bg        = PC_OPEN if self._flash_state else "#6b3000"
                    fg_sub    = PC_OPEN_FG
                    name_txt  = rec["name"][:12]
                    timer_txt = "● OPEN"
                    glow_state, glow_color = "open", PC_OPEN_FG
                    self._expire_timers.pop(i, None)

                elif rem > 300:
                    bg     = "#b91c1c" if self._flash_state else PC_BUSY
                    fg_sub = "#ffaaaa"
                    h = rem // 3600; m = (rem % 3600) // 60; s = rem % 60
                    name_txt  = rec["name"][:12]
                    timer_txt = f"{h}:{m:02d}:{s:02d}" if h else f"{m:02d}:{s:02d}"
                    glow_state, glow_color = "busy", PC_BUSY_FG
                    self._expire_timers.pop(i, None)

                elif rem > 0:
                    bg, fg_sub = PC_WARN, "#ffe0aa"
                    m = rem // 60; s = rem % 60
                    name_txt  = rec["name"][:12]
                    timer_txt = f"{m:02d}:{s:02d} ⚠"
                    glow_state, glow_color = "warn", PC_WARN_FG
                    self._expire_timers.pop(i, None)

                else:
                    # Expired — flash "TIME'S UP!" for 10 s then auto-free
                    if i not in self._expire_timers:
                        self._expire_timers[i] = now
                    elapsed = (now - self._expire_timers[i]).total_seconds()

                    if elapsed >= 10:
                        rec["session_type"]    = "closed"
                        del self._expire_timers[i]
                        need_persist           = True
                        bg, fg_sub             = PC_FREE, PC_FREE_FG
                        name_txt               = ""
                        timer_txt              = "Available"
                        glow_state, glow_color = "free", PC_FREE_FG
                    else:
                        bg        = PC_EXP_A if self._flash_state else PC_EXP_B
                        fg_sub    = "white"
                        name_txt  = rec["name"][:12] if rec else ""
                        timer_txt = "TIME'S UP!"
                        glow_state, glow_color = "expired", PC_EXP_A

            # Show BOOKED badge only when PC is occupied and has a pending booking
            has_booking = any(str(b.get("pc")) == str(i) and b.get("status") == "pending"
                              for b in self._bookings)
            booked_txt = "BOOKED" if (has_booking and rec is not None) else ""

            # Update canvas screen + text
            cv = info["canvas"]
            cv.itemconfig(info["screen"],      fill=bg)
            cv.itemconfig(info["text_name"],   fill=fg_sub, text=name_txt)
            cv.itemconfig(info["text_timer"],  fill=fg_sub, text=timer_txt)
            cv.itemconfig(info["text_booked"], text=booked_txt)
            info["glow_state"] = glow_state
            info["glow_color"] = glow_color

        # Update ON/OFF counter in header
        off_count = len(self._shutdown_pcs)
        self._pc_onoff_lbl.config(
            text=f"ON: {15 - off_count}  |  OFF: {off_count}")

        # Refresh form PC button colours (shutdown PCs go dark)
        try:
            self._highlight_form_pc(int(self._v_pc.get()))
        except (ValueError, AttributeError):
            pass

        if need_persist:
            self._persist()
            self._refresh_table()
            self._update_summary()

    # ── LEFT PANEL (FORM) ────────────────────────────────────────────────────
    def _build_form(self, parent):
        # Fixed-width panel — no scrolling needed
        panel = tk.Frame(parent, bg=BG_PANEL, width=382)
        panel.grid(row=0, column=0, sticky="ns", padx=(12, 6), pady=12)
        panel.grid_propagate(False)
        self._build_form_content(panel)

    def _build_form_content(self, p):
        """Build all form widgets inside the panel. No scrolling needed."""
        pad = {"padx": 12}

        # ── Declare ALL StringVars first so _recalc() is always safe ─────
        self._v_pc              = tk.StringVar(value="1")
        self._v_name            = tk.StringVar()
        self._v_dur_input       = tk.StringVar(value="1:00")
        self._v_timeout_display = tk.StringVar(value="—")
        self._v_dur             = tk.StringVar(value="—")
        self._v_amount          = tk.StringVar(value="0")
        self._v_disc            = tk.StringVar(value="0")
        self._v_final           = tk.StringVar(value="0")
        self._v_comment         = tk.StringVar()

        # ── PC Selector ───────────────────────────────────────────────────
        section_label(p, "SELECT PC")
        pc_grid = tk.Frame(p, bg=BG_PANEL)
        pc_grid.pack(fill="x", **pad, pady=(0, 2))
        for c in range(5):
            pc_grid.columnconfigure(c, weight=1)

        for i in range(1, 16):
            gc = (i - 1) % 5
            gr = (i - 1) // 5
            btn = tk.Button(pc_grid, text=str(i),
                            command=lambda n=i: self._select_form_pc(n),
                            bg=BG_CARD, fg=TEXT_MAIN, relief="flat",
                            font=("Segoe UI", 9, "bold"),
                            width=3, pady=3, cursor="hand2",
                            activebackground=ACCENT, activeforeground="white")
            btn.grid(row=gr, column=gc, padx=2, pady=2, sticky="ew")
            self._form_pc_btns[i] = btn

        self._highlight_form_pc(1)

        # ── Customer Name ─────────────────────────────────────────────────
        section_label(p, "CUSTOMER NAME")
        name_f = tk.Frame(p, bg=BG_CARD, highlightthickness=1,
                          highlightbackground=BORDER)
        name_f.pack(fill="x", **pad, pady=(0, 2))
        tk.Entry(name_f, textvariable=self._v_name, bg=BG_CARD, fg=TEXT_H,
                 insertbackground=TEXT_H, relief="flat",
                 font=("Segoe UI", 11), highlightthickness=0).pack(
            fill="x", padx=8, pady=5)

        # ── Time In ───────────────────────────────────────────────────────
        section_label(p, "TIME IN")
        self._tp_in = TimePicker(p)
        self._tp_in.pack(anchor="w", **pad, pady=(0, 2))
        self._tp_in.set_now()

        # ── Action Buttons (pinned to bottom of panel) ───────────────────
        # Pack with side="bottom" FIRST to always reserve space at the bottom.
        bottom_f = tk.Frame(p, bg=BG_PANEL)
        bottom_f.pack(side="bottom", fill="x")

        tk.Frame(bottom_f, bg=BORDER, height=1).pack(fill="x", padx=12, pady=(6, 5))

        self._btn_add = tk.Button(bottom_f, text="＋  ADD RECORD",
                                  command=self._add_or_update,
                                  bg=BTN_GREEN, fg="white",
                                  activebackground=BTN_GREEN_H,
                                  activeforeground="white",
                                  relief="flat",
                                  font=("Segoe UI", 11, "bold"),
                                  pady=10, cursor="hand2")
        self._btn_add.pack(fill="x", padx=12, pady=(0, 5))
        hover_bind(self._btn_add, BTN_GREEN, BTN_GREEN_H)

        self._btn_clear = tk.Button(bottom_f, text="✕  Clear Form",
                                    command=self._clear_form,
                                    bg=BG_CARD, fg=TEXT_DIM,
                                    activebackground=BG_MAIN,
                                    relief="flat", font=("Segoe UI", 9),
                                    pady=6, cursor="hand2",
                                    highlightthickness=1,
                                    highlightbackground=BORDER)
        self._btn_clear.pack(fill="x", padx=12, pady=(0, 10))

        # ── Dynamic middle section ────────────────────────────────────────
        # All three frames are children of `p` (packed after bottom_f is
        # reserved), managed by _update_form_visibility().

        # -- Duration section (hidden when editing_open) --
        self._dur_sec_frame = tk.Frame(p, bg=BG_PANEL)
        section_label(self._dur_sec_frame, "DURATION")
        dur_grid = tk.Frame(self._dur_sec_frame, bg=BG_PANEL)
        dur_grid.pack(fill="x", padx=12, pady=(0, 3))
        for c in range(4):
            dur_grid.columnconfigure(c, weight=1)

        # 8 preset buttons in rows 0 and 1
        for idx, (label, val) in enumerate(DURATION_PRESETS):
            btn = tk.Button(dur_grid, text=label,
                            command=lambda v=val: self._select_duration(v),
                            bg=BG_CARD, fg=TEXT_MAIN, relief="flat",
                            font=("Segoe UI", 9, "bold"),
                            pady=5, cursor="hand2",
                            activebackground=ACCENT2, activeforeground="white")
            btn.grid(row=idx // 4, column=idx % 4, padx=2, pady=2, sticky="ew")
            self._dur_btns[val] = btn

        # Open button — row 2, spans all 4 columns, orange
        open_btn = tk.Button(dur_grid, text="∞  Open",
                             command=lambda: self._select_duration("open"),
                             bg=BTN_ORANGE, fg="white", relief="flat",
                             font=("Segoe UI", 9, "bold"),
                             pady=5, cursor="hand2",
                             activebackground=BTN_ORANGE_H,
                             activeforeground="white")
        open_btn.grid(row=2, column=0, columnspan=4, padx=2, pady=2, sticky="ew")
        self._dur_btns["open"] = open_btn

        # Custom duration entry (hidden when "open" selected)
        self._custom_dur_frame = tk.Frame(self._dur_sec_frame, bg=BG_PANEL)
        tk.Label(self._custom_dur_frame, text="Custom:", bg=BG_PANEL, fg=TEXT_DIM,
                 font=("Segoe UI", 8)).pack(side="left", padx=(0, 4))
        cust_inp_f = tk.Frame(self._custom_dur_frame, bg=BG_CARD,
                              highlightthickness=1, highlightbackground=BORDER)
        cust_inp_f.pack(side="left", fill="x", expand=True)
        self._custom_dur_entry = tk.Entry(
            cust_inp_f, textvariable=self._v_dur_input,
            bg=BG_CARD, fg=TEXT_H, insertbackground=TEXT_H,
            relief="flat", font=("Consolas", 10), highlightthickness=0, width=8)
        self._custom_dur_entry.pack(side="left", padx=6, pady=4)
        tk.Label(cust_inp_f, text="(e.g. 1:15)", bg=BG_CARD, fg=TEXT_DIM,
                 font=("Segoe UI", 7)).pack(side="left", padx=(0, 6))
        self._custom_dur_entry.bind("<FocusOut>", lambda e: self._on_custom_dur())
        self._custom_dur_entry.bind("<Return>",   lambda e: self._on_custom_dur())

        # Time-out display label (hidden when "open" selected)
        self._tout_display_frame = tk.Frame(self._dur_sec_frame, bg=BG_CARD,
                                            highlightthickness=1,
                                            highlightbackground=BORDER)
        tk.Label(self._tout_display_frame, text="Time Out:", bg=BG_CARD,
                 fg=TEXT_DIM, font=("Segoe UI", 8)).pack(
            side="left", padx=(8, 4), pady=5)
        tk.Label(self._tout_display_frame, textvariable=self._v_timeout_display,
                 bg=BG_CARD, fg=ACCENT, font=("Consolas", 11, "bold")).pack(
            side="left", pady=5)
        tk.Label(self._tout_display_frame, textvariable=self._v_dur,
                 bg=BG_CARD, fg=TEXT_DIM, font=("Segoe UI", 8)).pack(
            side="right", padx=8)

        # -- Time Out picker (only shown when editing_open) --
        self._tout_picker_frame = tk.Frame(p, bg=BG_PANEL)
        section_label(self._tout_picker_frame, "TIME OUT")
        self._tp_out = TimePicker(self._tout_picker_frame)
        self._tp_out.pack(anchor="w", padx=12, pady=(0, 2))
        for sp in (self._tp_out._sh, self._tp_out._sm, self._tp_out._sa):
            sp.bind("<FocusOut>",           lambda e: self._recalc_open_edit(), add='+')
            sp.bind("<<ComboboxSelected>>", lambda e: self._recalc_open_edit(), add='+')
            sp.bind("<ButtonRelease>",      lambda e: self.after(100, self._recalc_open_edit), add='+')

        # -- Payment section (Amount only) --
        self._payment_sec_frame = tk.Frame(p, bg=BG_PANEL)
        section_label(self._payment_sec_frame, "PAYMENT")
        amt_row_f = tk.Frame(self._payment_sec_frame, bg=BG_PANEL)
        amt_row_f.pack(fill="x", padx=12, pady=2)
        tk.Label(amt_row_f, text="Amount", bg=BG_PANEL, fg=TEXT_DIM,
                 font=("Segoe UI", 8), width=9, anchor="w").pack(side="left")
        amt_inp_f = tk.Frame(amt_row_f, bg=BG_CARD, highlightthickness=1,
                             highlightbackground=BORDER)
        amt_inp_f.pack(side="left", fill="x", expand=True)
        tk.Label(amt_inp_f, text="₱", bg=BG_CARD, fg=ACCENT,
                 font=("Consolas", 16, "bold")).pack(side="left", padx=(6, 0))
        self._amt_entry = tk.Entry(
            amt_inp_f, textvariable=self._v_amount, bg=BG_CARD, fg=TEXT_H,
            insertbackground=TEXT_H, relief="flat",
            font=("Consolas", 16, "bold"), highlightthickness=0)
        self._amt_entry.pack(side="left", fill="x", expand=True, padx=4, pady=10)
        self._amt_entry.bind("<Key>", lambda e: setattr(self, "_amount_manual", True))
        self._v_amount.trace_add("write", self._on_amount_changed)

        # Comment (optional)
        cmt_row_f = tk.Frame(self._payment_sec_frame, bg=BG_PANEL)
        cmt_row_f.pack(fill="x", padx=12, pady=(0, 4))
        tk.Label(cmt_row_f, text="Comment", bg=BG_PANEL, fg=TEXT_DIM,
                 font=("Segoe UI", 8), width=9, anchor="w").pack(side="left")
        cmt_inp_f = tk.Frame(cmt_row_f, bg=BG_CARD, highlightthickness=1,
                             highlightbackground=BORDER)
        cmt_inp_f.pack(side="left", fill="x", expand=True)
        tk.Entry(cmt_inp_f, textvariable=self._v_comment, bg=BG_CARD, fg=TEXT_DIM,
                 insertbackground=TEXT_H, relief="flat",
                 font=("Segoe UI", 9), highlightthickness=0).pack(
            fill="x", padx=8, pady=5)

        # ── bind Time-In spinners ─────────────────────────────────────────
        for sp in (self._tp_in._sh, self._tp_in._sm, self._tp_in._sa):
            sp.bind("<FocusOut>",           lambda e: self._recalc(), add='+')
            sp.bind("<<ComboboxSelected>>", lambda e: self._recalc(), add='+')
            sp.bind("<ButtonRelease>",      lambda e: self.after(100, self._recalc), add='+')

        # Show initial state (normal fixed session)
        self._update_form_visibility()
        self.after(80, self._recalc)

    def _highlight_form_pc(self, pc_num):
        for n, btn in self._form_pc_btns.items():
            if n == pc_num:
                btn.config(bg=ACCENT, fg="white")
            elif n in self._shutdown_pcs:
                btn.config(bg="#1a1a1a", fg="#404040")
            else:
                btn.config(bg=BG_CARD, fg=TEXT_MAIN)

    def _update_form_visibility(self):
        """Show/hide dynamic form sections based on current state."""
        # Hide all dynamic sections first
        self._dur_sec_frame.pack_forget()
        self._tout_picker_frame.pack_forget()
        self._payment_sec_frame.pack_forget()

        if self._editing_open:
            # Editing an existing OPEN record: show Time Out picker + Amount
            self._tout_picker_frame.pack(fill="x")
            self._payment_sec_frame.pack(fill="x")
            self._btn_add.config(text="💾  CLOSE & SAVE",
                                 bg=BTN_GREEN, activebackground=BTN_GREEN_H)
            hover_bind(self._btn_add, BTN_GREEN, BTN_GREEN_H)
        elif self._selected_dur == "open":
            # New open session: duration grid + amount (no custom dur / tout display)
            self._dur_sec_frame.pack(fill="x")
            self._custom_dur_frame.pack_forget()
            self._tout_display_frame.pack_forget()
            self._payment_sec_frame.pack(fill="x")
            self._btn_add.config(text="▶  START SESSION",
                                 bg=BTN_ORANGE, activebackground=BTN_ORANGE_H)
            hover_bind(self._btn_add, BTN_ORANGE, BTN_ORANGE_H)
        else:
            # Normal fixed session
            self._dur_sec_frame.pack(fill="x")
            self._custom_dur_frame.pack(fill="x", padx=12, pady=(0, 2))
            self._tout_display_frame.pack(fill="x", padx=12, pady=(0, 2))
            self._payment_sec_frame.pack(fill="x")
            if self._edit_idx is not None:
                self._btn_add.config(text="💾  SAVE CHANGES",
                                     bg=ACCENT2, activebackground=ACCENT2_H)
                hover_bind(self._btn_add, ACCENT2, ACCENT2_H)
            else:
                self._btn_add.config(text="＋  ADD RECORD",
                                     bg=BTN_GREEN, activebackground=BTN_GREEN_H)
                hover_bind(self._btn_add, BTN_GREEN, BTN_GREEN_H)
        # Disable button if selected PC is shutdown (new records only)
        self._update_add_btn_state()

    def _recalc_open_edit(self):
        """Auto-calculate amount from Time In / Time Out (editing-open mode)."""
        m_in  = self._tp_in.get_minutes()
        m_out = self._tp_out.get_minutes()
        diff  = m_out - m_in
        if diff < 0:
            diff += 24 * 60
        if not self._amount_manual:
            self._v_amount.set(str(int(calc_amount(diff, self._cfg))))

    def _select_form_pc(self, pc_num):
        self._v_pc.set(str(pc_num))
        self._highlight_form_pc(pc_num)
        self._update_add_btn_state()

    def _select_duration(self, val):
        self._selected_dur = val
        # Update button highlight colours
        for v, btn in self._dur_btns.items():
            if v == "open":
                btn.config(bg=BTN_ORANGE_H if v == val else BTN_ORANGE, fg="white")
            else:
                btn.config(bg=ACCENT2 if v == val else BG_CARD,
                           fg="white"  if v == val else TEXT_MAIN)
        if val == "open":
            self._v_amount.set("0")
            self._amount_manual = False
        else:
            self._v_dur_input.set(val)
            self._amount_manual = False   # duration click always refreshes amount
            self._recalc()
        self._update_form_visibility()

    def _on_custom_dur(self):
        """Called when user edits the custom duration entry manually."""
        # De-highlight all preset buttons (typed value may not match any)
        typed = self._v_dur_input.get().strip()
        for v, btn in self._dur_btns.items():
            btn.config(bg=ACCENT2 if v == typed else BG_CARD,
                       fg="white"  if v == typed else TEXT_MAIN)
        self._recalc()

    # ── TABLE PANEL ─────────────────────────────────────────────────────────
    def _build_table(self, parent):
        frame = tk.Frame(parent, bg=BG_MAIN)
        frame.grid(row=0, column=1, sticky="nsew", padx=(6, 12), pady=12)
        frame.columnconfigure(0, weight=1)
        frame.rowconfigure(2, weight=1)   # row 2 = treeviews

        # ── header row ───────────────────────────────────────────────────────
        thead = tk.Frame(frame, bg=BG_MAIN)
        thead.grid(row=0, column=0, columnspan=2, sticky="ew", pady=(0, 4))
        thead.columnconfigure(0, weight=1)

        tk.Label(thead, text="TODAY'S SESSIONS", bg=BG_MAIN, fg=TEXT_H,
                 font=("Segoe UI", 11, "bold")).grid(row=0, column=0, sticky="w")

        btn_bar = tk.Frame(thead, bg=BG_MAIN)
        btn_bar.grid(row=0, column=1, sticky="e")

        edit_btn = tk.Button(btn_bar, text="✏  Edit Selected",
                             command=self._edit_selected,
                             bg=BG_CARD, fg=TEXT_MAIN,
                             activebackground=ACCENT2, activeforeground="white",
                             relief="flat", font=("Segoe UI", 9),
                             padx=10, pady=5, cursor="hand2",
                             highlightthickness=1, highlightbackground=BORDER)
        edit_btn.pack(side="left", padx=(0, 6))
        hover_bind(edit_btn, BG_CARD, ACCENT2)

        folder_btn = tk.Button(btn_bar, text="📂  Excel Folder",
                               command=self._open_folder,
                               bg=BG_CARD, fg=TEXT_DIM,
                               activebackground=BG_PANEL,
                               relief="flat", font=("Segoe UI", 9),
                               padx=10, pady=5, cursor="hand2",
                               highlightthickness=1, highlightbackground=BORDER)
        folder_btn.pack(side="left")

        # ── tab bar ───────────────────────────────────────────────────────────
        tab_bar = tk.Frame(frame, bg=BG_MAIN)
        tab_bar.grid(row=1, column=0, columnspan=2, sticky="ew", pady=(0, 0))

        self._tab_btns = {}
        for tab_id, tab_label in (("current",  "▶  Current Users"),
                                   ("all",      "☰  All Records"),
                                   ("expenses", "💰  Expenses"),
                                   ("bookings", "🔖  Bookings")):
            btn = tk.Button(tab_bar, text=tab_label,
                            command=lambda t=tab_id: self._switch_tab(t),
                            bg=BG_CARD, fg=TEXT_DIM,
                            activebackground=ACCENT2, activeforeground="white",
                            relief="flat", font=("Segoe UI", 9, "bold"),
                            padx=14, pady=7, cursor="hand2",
                            highlightthickness=0, borderwidth=0)
            btn.pack(side="left", padx=(0, 3))
            self._tab_btns[tab_id] = btn

        # thin accent line below tabs
        tk.Frame(frame, bg=ACCENT2, height=2).grid(
            row=1, column=0, columnspan=2, sticky="sew")

        # ── shared treeview style ─────────────────────────────────────────────
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Pro.Treeview",
                         background=BG_MAIN, foreground=TEXT_MAIN,
                         fieldbackground=BG_MAIN,
                         rowheight=40, borderwidth=0,
                         font=("Segoe UI", 16))
        style.configure("Pro.Treeview.Heading",
                         background=BG_CARD, foreground=TEXT_DIM,
                         font=("Segoe UI", 8, "bold"), relief="flat",
                         padding=(6, 6))
        style.map("Pro.Treeview",
                  background=[("selected", ROW_SEL)],
                  foreground=[("selected", TEXT_H)])

        cols   = ("pc", "name", "time_in", "time_out", "duration", "amount", "comment")
        hdrs   = ("PC", "Name", "Time In", "Time Out", "Duration", "Amount", "Comment")
        widths = (44,   155,    90,        90,         85,         80,       140)

        def _make_tree():
            tv = ttk.Treeview(frame, columns=cols, show="headings",
                              style="Pro.Treeview", selectmode="browse")
            for col, hdr, w in zip(cols, hdrs, widths):
                tv.heading(col, text=hdr)
                tv.column(col, width=w, anchor="center", minwidth=w)
            tv.column("name",    anchor="w")
            tv.column("comment", anchor="w")
            tv.tag_configure("odd",    background=ROW_ODD,    foreground=TEXT_MAIN)
            tv.tag_configure("even",   background=ROW_EVEN,   foreground=TEXT_MAIN)
            tv.tag_configure("open",   background=ROW_ODD,    foreground=TEXT_MAIN)
            tv.tag_configure("closed", background=ROW_ODD,    foreground=TEXT_DIM)
            tv.bind("<Double-1>", lambda e: self._edit_selected())
            return tv

        # ── Current Users tree (row=2) ────────────────────────────────────────
        self._tree_cur = _make_tree()
        self._vsb_cur  = ttk.Scrollbar(frame, orient="vertical",
                                        command=self._tree_cur.yview)
        self._tree_cur.configure(yscrollcommand=self._vsb_cur.set)
        self._tree_cur.grid(row=2, column=0, sticky="nsew")
        self._vsb_cur.grid(row=2, column=1, sticky="ns")

        # ── All Records tree (same row, hidden initially) ─────────────────────
        self._tree_all = _make_tree()
        self._vsb_all  = ttk.Scrollbar(frame, orient="vertical",
                                        command=self._tree_all.yview)
        self._tree_all.configure(yscrollcommand=self._vsb_all.set)
        self._tree_all.grid(row=2, column=0, sticky="nsew")
        self._vsb_all.grid(row=2, column=1, sticky="ns")

        # ── Expenses panel (same row=2, hidden initially) ─────────────────────
        self._exp_frame = tk.Frame(frame, bg=BG_MAIN)
        self._exp_frame.grid(row=2, column=0, columnspan=2, sticky="nsew")
        self._build_expenses_tab(self._exp_frame)

        # ── Bookings panel (same row=2, hidden initially) ──────────────────────
        self._bk_frame = tk.Frame(frame, bg=BG_MAIN)
        self._bk_frame.grid(row=2, column=0, columnspan=2, sticky="nsew")
        self._build_bookings_tab(self._bk_frame)

        # activate Current Users tab by default
        self._active_tab  = "current"
        self._active_tree = self._tree_cur
        self._switch_tab("current")

    def _build_expenses_tab(self, parent):
        parent.columnconfigure(0, weight=1)
        parent.rowconfigure(2, weight=1)

        # ── Summary card ─────────────────────────────────────────────────────
        card = tk.Frame(parent, bg=BG_CARD, highlightthickness=1,
                        highlightbackground=BORDER)
        card.grid(row=0, column=0, sticky="ew", padx=0, pady=(0, 1))
        card.columnconfigure(1, weight=1)

        tk.Frame(card, bg=BTN_GREEN, width=4).grid(
            row=0, column=0, rowspan=3, sticky="ns")

        lbl_kw = dict(bg=BG_CARD, font=("Segoe UI", 8, "bold"), anchor="w")
        val_kw = dict(bg=BG_CARD, font=("Consolas", 11, "bold"))

        tk.Label(card, text="EARNINGS",  fg=TEXT_DIM,  **lbl_kw).grid(
            row=0, column=1, padx=(12, 4), pady=(8, 1), sticky="w")
        self._exp_earn_lbl = tk.Label(card, text="₱ 0", fg=BTN_GREEN_H, **val_kw)
        self._exp_earn_lbl.grid(row=0, column=2, padx=(0, 16), pady=(8, 1), sticky="e")

        tk.Label(card, text="EXPENSES",  fg=TEXT_DIM,  **lbl_kw).grid(
            row=1, column=1, padx=(12, 4), pady=1, sticky="w")
        self._exp_total_lbl = tk.Label(card, text="₱ 0", fg=ACCENT, **val_kw)
        self._exp_total_lbl.grid(row=1, column=2, padx=(0, 16), pady=1, sticky="e")

        tk.Label(card, text="REMAINING", fg=TEXT_DIM,  **lbl_kw).grid(
            row=2, column=1, padx=(12, 4), pady=(1, 8), sticky="w")
        self._exp_rem_lbl = tk.Label(card, text="₱ 0", fg=ACCENT2_H, **val_kw)
        self._exp_rem_lbl.grid(row=2, column=2, padx=(0, 16), pady=(1, 8), sticky="e")

        # ── Add Expense form ─────────────────────────────────────────────────
        form = tk.Frame(parent, bg=BG_PANEL)
        form.grid(row=1, column=0, sticky="ew", pady=(0, 1))
        form.columnconfigure(1, weight=1)
        form.columnconfigure(3, weight=1)

        tk.Label(form, text="Name", bg=BG_PANEL, fg=TEXT_DIM,
                 font=("Segoe UI", 8)).grid(row=0, column=0, padx=(12, 4),
                                             pady=8, sticky="w")
        self._v_exp_name = tk.StringVar()
        name_f = tk.Frame(form, bg=BG_CARD, highlightthickness=1,
                          highlightbackground=BORDER)
        name_f.grid(row=0, column=1, sticky="ew", padx=(0, 12), pady=6)
        tk.Entry(name_f, textvariable=self._v_exp_name, bg=BG_CARD, fg=TEXT_H,
                 insertbackground=TEXT_H, relief="flat",
                 font=("Segoe UI", 10), highlightthickness=0).pack(
            fill="x", padx=6, pady=4)

        tk.Label(form, text="Amount", bg=BG_PANEL, fg=TEXT_DIM,
                 font=("Segoe UI", 8)).grid(row=0, column=2, padx=(0, 4),
                                             pady=8, sticky="w")
        self._v_exp_amount = tk.StringVar()
        amt_f = tk.Frame(form, bg=BG_CARD, highlightthickness=1,
                         highlightbackground=BORDER)
        amt_f.grid(row=0, column=3, sticky="ew", padx=(0, 12), pady=6)
        tk.Label(amt_f, text="₱", bg=BG_CARD, fg=ACCENT,
                 font=("Consolas", 10, "bold")).pack(side="left", padx=(5, 0))
        tk.Entry(amt_f, textvariable=self._v_exp_amount, bg=BG_CARD, fg=TEXT_H,
                 insertbackground=TEXT_H, relief="flat",
                 font=("Consolas", 10), highlightthickness=0, width=8).pack(
            side="left", padx=4, pady=4)

        add_btn = tk.Button(form, text="＋  Add Expense",
                            command=self._add_expense,
                            bg=BTN_GREEN, fg="white",
                            activebackground=BTN_GREEN_H,
                            relief="flat", font=("Segoe UI", 9, "bold"),
                            padx=12, pady=6, cursor="hand2")
        add_btn.grid(row=0, column=4, padx=(0, 12), pady=6)
        hover_bind(add_btn, BTN_GREEN, BTN_GREEN_H)

        # ── Expense list with per-row Edit / Delete ───────────────────────────
        list_hdr = tk.Frame(parent, bg=BG_MAIN)
        list_hdr.grid(row=2, column=0, sticky="nsew")
        list_hdr.columnconfigure(0, weight=1)
        list_hdr.rowconfigure(1, weight=1)

        hdr_row = tk.Frame(list_hdr, bg=BG_MAIN)
        hdr_row.grid(row=0, column=0, sticky="ew", pady=(4, 2))
        tk.Label(hdr_row, text="EXPENSE LIST", bg=BG_MAIN, fg=TEXT_DIM,
                 font=("Segoe UI", 8, "bold")).pack(side="left", padx=12)

        # Scrollable canvas + inner frame for rows
        scrl_f = tk.Frame(list_hdr, bg=BG_MAIN)
        scrl_f.grid(row=1, column=0, sticky="nsew")
        scrl_f.columnconfigure(0, weight=1)
        scrl_f.rowconfigure(0, weight=1)

        self._exp_canvas = tk.Canvas(scrl_f, bg=BG_MAIN, highlightthickness=0)
        vsb_exp = ttk.Scrollbar(scrl_f, orient="vertical",
                                command=self._exp_canvas.yview)
        self._exp_canvas.configure(yscrollcommand=vsb_exp.set)
        self._exp_canvas.grid(row=0, column=0, sticky="nsew")
        vsb_exp.grid(row=0, column=1, sticky="ns")

        self._exp_list_frame = tk.Frame(self._exp_canvas, bg=BG_MAIN)
        self._exp_canvas_win = self._exp_canvas.create_window(
            (0, 0), window=self._exp_list_frame, anchor="nw")

        self._exp_list_frame.bind(
            "<Configure>",
            lambda e: self._exp_canvas.configure(
                scrollregion=self._exp_canvas.bbox("all")))
        self._exp_canvas.bind(
            "<Configure>",
            lambda e: self._exp_canvas.itemconfig(
                self._exp_canvas_win, width=e.width))

    def _switch_tab(self, tab_id):
        self._active_tab = tab_id
        for tid, btn in self._tab_btns.items():
            if tid == tab_id:
                btn.config(bg=ACCENT2, fg="white")
            else:
                btn.config(bg=BG_CARD, fg=TEXT_DIM)

        # hide everything first
        self._tree_cur.grid_remove(); self._vsb_cur.grid_remove()
        self._tree_all.grid_remove(); self._vsb_all.grid_remove()
        self._exp_frame.grid_remove()
        self._bk_frame.grid_remove()

        if tab_id == "current":
            self._active_tree = self._tree_cur
            self._tree_cur.grid(); self._vsb_cur.grid()
        elif tab_id == "all":
            self._active_tree = self._tree_all
            self._tree_all.grid(); self._vsb_all.grid()
        elif tab_id == "expenses":
            self._active_tree = None
            self._exp_frame.grid()
            self._refresh_expense_summary()
        else:  # bookings
            self._active_tree = None
            self._bk_frame.grid()
            self._refresh_bookings()
            self._refresh_bk_pc_btns()

    # ── SUMMARY BAR ─────────────────────────────────────────────────────────
    def _build_summary(self):
        bar = tk.Frame(self, bg=BG_CARD)
        bar.grid(row=3, column=0, sticky="ew")
        bar.columnconfigure(4, weight=1)   # notification area expands

        # left accent strip
        tk.Frame(bar, bg=ACCENT, width=4).grid(row=0, column=0, sticky="ns")

        tk.Label(bar, text="TODAY'S TOTAL", bg=BG_CARD, fg=TEXT_DIM,
                 font=("Segoe UI", 8, "bold")).grid(
            row=0, column=1, padx=(16, 8), pady=12)

        self._total_lbl = tk.Label(bar, text="₱ 0.00", bg=BG_CARD, fg=TEXT_H,
                                   font=("Segoe UI", 20, "bold"))
        self._total_lbl.grid(row=0, column=2, sticky="w")

        self._sessions_lbl = tk.Label(bar, text="0 sessions", bg=BG_CARD,
                                      fg=TEXT_DIM, font=("Segoe UI", 9))
        self._sessions_lbl.grid(row=0, column=3, padx=20)

        # Divider
        tk.Frame(bar, bg=BORDER, width=1).grid(row=0, column=4, sticky="ns",
                                               padx=(0, 12), pady=8)

        # Notification label — right side of bar
        self._notif_lbl = tk.Label(bar, text="", bg=BG_CARD, fg="#fb923c",
                                   font=("Segoe UI", 9, "bold"),
                                   anchor="w", justify="left")
        self._notif_lbl.grid(row=0, column=5, sticky="ew", padx=(0, 16))

        # Start rotation ticker
        self.after(3000, self._rotate_notification)

    # ── NOTIFICATION BAR ─────────────────────────────────────────────────────
    def _push_notification(self, text, color, key):
        """Add or refresh a notification. `key` prevents duplicates."""
        # Remove existing entry with same key, then append fresh
        self._notifications = [n for n in self._notifications if n["key"] != key]
        self._notifications.append({"text": text, "color": color, "key": key})
        self._notif_index = len(self._notifications) - 1
        self._show_current_notification()

    def _remove_notification(self, key):
        self._notifications = [n for n in self._notifications if n["key"] != key]
        if self._notifications:
            self._notif_index = self._notif_index % len(self._notifications)
        else:
            self._notif_index = 0
        self._show_current_notification()

    def _show_current_notification(self):
        if not self._notifications:
            self._notif_lbl.config(text="", fg="#fb923c")
            return
        n     = self._notifications[self._notif_index % len(self._notifications)]
        count = len(self._notifications)
        prefix = f"🔔 ({count}) " if count > 1 else "🔔 "
        self._notif_lbl.config(text=prefix + n["text"], fg=n["color"])

    def _rotate_notification(self):
        if len(self._notifications) > 1:
            self._notif_index = (self._notif_index + 1) % len(self._notifications)
            self._show_current_notification()
        self.after(3000, self._rotate_notification)

    def _sync_notifications(self):
        """Remove stale notifications whose session/booking is gone."""
        now    = datetime.now()
        active_ids = set()
        for rec in self._records:
            rid = rec.get("id", "")
            st  = rec.get("session_type", "fixed")
            if st == "closed":
                continue
            t_out = parse_session_time(rec.get("time_out", ""))
            if t_out and t_out > now:
                active_ids.add(f"warn_{rid}")
            active_ids.add(f"exp_{rid}")   # keep expired until actually closed

        pending_bk_ids = {f"bk_{b['id']}" for b in self._bookings
                          if b.get("status") == "pending"}
        active_ids |= pending_bk_ids

        stale = [n["key"] for n in self._notifications
                 if n["key"] not in active_ids]
        for key in stale:
            self._remove_notification(key)

    # ── RECALCULATE ──────────────────────────────────────────────────────────
    def _recalc(self):
        _, m_in  = self._tp_in.get_time()
        dur_str  = self._v_dur_input.get().strip()
        dur_min  = parse_dur_input(dur_str)
        if dur_min > 0:
            self._v_timeout_display.set(calc_timeout_str(m_in, dur_min))
        else:
            self._v_timeout_display.set("—")
        self._v_dur.set(fmt_duration(dur_min))
        if not self._amount_manual:
            rate_key = f"rate_{dur_str}"
            if rate_key in self._cfg:
                self._v_amount.set(str(int(self._cfg[rate_key])))
            else:
                self._v_amount.set(str(int(calc_amount(dur_min, self._cfg))))
        self._update_final()

    def _on_amount_changed(self, *_): self._update_final()
    def _on_disc_changed(self, *_):   self._update_final()

    def _update_final(self):
        try:    amt  = float(self._v_amount.get() or 0)
        except: amt  = 0
        try:    disc = float(self._v_disc.get() or 0)
        except: disc = 0
        final = max(0, amt - disc)
        self._v_final.set(f"{final:.0f}")
        # colour the total label green if final > 0
        self._v_final.set(str(int(final)))

    # ── ADD / UPDATE ─────────────────────────────────────────────────────────
    def _add_or_update(self):
        name = self._v_name.get().strip()
        if not name:
            messagebox.showwarning("Missing", "Please enter a customer name.")
            return

        if self._editing_open:
            # ── Closing an existing OPEN session ──────────────────────────
            try:
                amt = float(self._v_amount.get() or 0)
            except ValueError:
                messagebox.showerror("Error", "Amount must be a number.")
                return
            m_in  = self._tp_in.get_minutes()
            m_out = self._tp_out.get_minutes()
            diff  = m_out - m_in
            if diff < 0:
                diff += 24 * 60
            old_rec = dict(self._records[self._edit_idx])
            rec = dict(self._records[self._edit_idx])
            rec["name"]         = name
            rec["time_out"]     = self._tp_out.get_str()
            rec["duration"]     = fmt_duration(diff)
            rec["amount"]       = amt
            rec["discount"]     = 0
            rec["final"]        = amt
            rec["session_type"] = "fixed"
            rec["comment"]      = self._v_comment.get().strip()[:100]
            self._records[self._edit_idx] = rec
            self._log_edit(old_rec, rec)
            self._edit_idx     = None
            self._editing_open = False

        elif self._selected_dur == "open":
            # ── Starting a new OPEN session ───────────────────────────────
            pc     = self._v_pc.get()
            active = self._get_active_session(pc)
            if active:
                messagebox.showwarning(
                    "PC In Use",
                    f"PC {pc} is already in use by {active['name']}!\n\n"
                    "Wait until their session ends or edit the existing record."
                )
                return
            rec = {
                "pc":           pc,
                "name":         name,
                "time_in":      self._tp_in.get_str(),
                "time_out":     "OPEN",
                "duration":     "OPEN",
                "amount":       0,
                "discount":     0,
                "final":        0,
                "session_type": "open",
                "advance":      0,
                "comment":      self._v_comment.get().strip()[:100],
            }
            rec["id"] = next_record_id(self._records)
            self._records.append(rec)

        else:
            # ── Normal fixed session (new or edit) ────────────────────────
            if self._edit_idx is None:
                pc     = self._v_pc.get()
                active = self._get_active_session(pc)
                if active:
                    messagebox.showwarning(
                        "PC In Use",
                        f"PC {pc} is already in use by {active['name']}!\n\n"
                        "Wait until their session ends or edit the existing record."
                    )
                    return
            try:
                amt = float(self._v_amount.get() or 0)
            except ValueError:
                messagebox.showerror("Error", "Amount must be a number.")
                return
            rec = {
                "pc":           self._v_pc.get(),
                "name":         name,
                "time_in":      self._tp_in.get_str(),
                "time_out":     self._v_timeout_display.get(),
                "duration":     self._v_dur.get(),
                "amount":       amt,
                "discount":     0,
                "final":        amt,
                "session_type": "fixed",
                "advance":      0,
                "comment":      self._v_comment.get().strip()[:100],
            }
            if self._edit_idx is not None:
                old_rec = dict(self._records[self._edit_idx])
                rec["id"] = self._records[self._edit_idx].get("id", "")
                self._records[self._edit_idx] = rec
                self._log_edit(old_rec, rec)
                self._edit_idx = None
            else:
                rec["id"] = next_record_id(self._records)
                self._records.append(rec)

        self._persist()
        self._refresh_table()
        self._update_summary()
        self._update_pc_grid()
        self._sync_notifications()
        self._clear_form()

    def _log_edit(self, old_rec, new_rec):
        """Compare old and new record dicts; append any changes to self._edit_log."""
        _FIELDS = [("name", "Name"), ("time_in", "Time In"),
                   ("time_out", "Time Out"), ("duration", "Duration"),
                   ("amount", "Amount"), ("comment", "Comment")]
        parts = []
        for key, label in _FIELDS:
            ov, nv = str(old_rec.get(key, "")), str(new_rec.get(key, ""))
            if ov != nv:
                parts.append(f"{label} changed from {ov} to {nv}")
        if parts:
            self._edit_log.append({
                "timestamp": datetime.now().strftime("%H:%M:%S"),
                "pc":        str(new_rec.get("pc", "")),
                "name":      new_rec.get("name", ""),
                "summary":   "; ".join(parts),
                "changes":   [{"field": p.split(" changed")[0],
                               "old":   p.split(" from ")[1].split(" to ")[0],
                               "new":   p.split(" to ")[-1]}
                              for p in parts],
            })

    def _persist(self):
        save_cafe_data(self._records, self._expenses, self._shutdown_pcs,
                       self._bookings, self._edit_log)
        try:
            save_to_excel(self._records, self._expenses)
        except Exception as ex:
            messagebox.showerror("Excel Error", str(ex))

    # ── TABLE HELPERS ────────────────────────────────────────────────────────
    def _refresh_table(self):
        self._tree_cur.delete(*self._tree_cur.get_children())
        self._tree_all.delete(*self._tree_all.get_children())

        cur_row = 0   # separate row counter so alternating colours are tight
        for i, rec in enumerate(self._records):
            is_open = (rec.get("session_type") == "open"
                       and rec.get("time_out") == "OPEN")
            t_out    = "" if is_open else rec["time_out"]
            duration = "" if is_open else rec["duration"]
            amount   = ("" if is_open and rec["amount"] == 0
                        else f"₱{rec['amount']:.0f}")
            vals = (rec["pc"], rec["name"],
                    rec["time_in"], t_out,
                    duration, amount,
                    rec.get("comment", ""))
            st = rec.get("session_type", "fixed")

            # All Records — every session ever, always
            all_tag = ("open"   if st == "open"   else
                       "closed" if st == "closed"  else
                       "odd"    if i % 2 == 0      else "even")
            self._tree_all.insert("", "end", iid=str(i), tags=(all_tag,),
                                  values=vals)

            # Current Users — only sessions not yet closed
            if st != "closed":
                cur_tag = "open" if st == "open" else (
                          "odd"  if cur_row % 2 == 0 else "even")
                self._tree_cur.insert("", "end", iid=str(i), tags=(cur_tag,),
                                      values=vals)
                cur_row += 1

    def _update_summary(self):
        total = sum(r["final"] for r in self._records)
        self._total_lbl.config(text=f"₱ {total:,.2f}")
        n = len(self._records)
        self._sessions_lbl.config(text=f"{n} session{'s' if n != 1 else ''}")
        self._refresh_expense_summary()

    def _refresh_expense_summary(self):
        earnings  = sum(r["final"] for r in self._records)
        exp_total = sum(e["amount"] for e in self._expenses)
        remaining = max(0, earnings - exp_total)
        self._exp_earn_lbl.config( text=f"₱ {earnings:,.0f}")
        self._exp_total_lbl.config(text=f"₱ {exp_total:,.0f}")
        self._exp_rem_lbl.config(  text=f"₱ {remaining:,.0f}")

    def _refresh_expenses(self):
        for w in self._exp_list_frame.winfo_children():
            w.destroy()
        for i, exp in enumerate(self._expenses):
            bg = ROW_ODD if i % 2 == 0 else ROW_EVEN
            row = tk.Frame(self._exp_list_frame, bg=bg)
            row.pack(fill="x")
            tk.Label(row, text=exp["name"], bg=bg, fg=TEXT_MAIN,
                     font=("Segoe UI", 9), anchor="w").pack(
                side="left", padx=(12, 4), pady=6, fill="x", expand=True)
            tk.Label(row, text=f"₱ {exp['amount']:,.0f}", bg=bg, fg=TEXT_H,
                     font=("Consolas", 9, "bold"), width=10, anchor="e").pack(
                side="left", padx=(0, 8), pady=6)
            edit_btn = tk.Button(row, text="✏",
                                 command=lambda idx=i: self._edit_expense(idx),
                                 bg=BG_CARD, fg=ACCENT2, relief="flat",
                                 font=("Segoe UI", 8), padx=6, pady=3,
                                 cursor="hand2", highlightthickness=1,
                                 highlightbackground=BORDER)
            edit_btn.pack(side="left", padx=(0, 4), pady=4)
            hover_bind(edit_btn, BG_CARD, ACCENT2)
            del_btn = tk.Button(row, text="✕",
                                command=lambda idx=i: self._delete_expense_at(idx),
                                bg=BG_CARD, fg=ACCENT, relief="flat",
                                font=("Segoe UI", 8), padx=6, pady=3,
                                cursor="hand2", highlightthickness=1,
                                highlightbackground=BORDER)
            del_btn.pack(side="left", padx=(0, 12), pady=4)
            hover_bind(del_btn, BG_CARD, ACCENT)
        self._refresh_expense_summary()

    def _edit_expense(self, idx):
        exp = self._expenses[idx]
        dlg = tk.Toplevel(self)
        dlg.title("Edit Expense")
        dlg.configure(bg=BG_MAIN)
        dlg.resizable(False, False)
        dlg.grab_set()
        dlg.geometry("+%d+%d" % (self.winfo_rootx() + 200,
                                  self.winfo_rooty() + 200))

        # Header
        hdr = tk.Frame(dlg, bg=BG_CARD)
        hdr.pack(fill="x")
        tk.Frame(hdr, bg=ACCENT2, width=4).pack(side="left", fill="y")
        tk.Label(hdr, text="Edit Expense", bg=BG_CARD, fg=TEXT_H,
                 font=("Segoe UI", 11, "bold"), padx=14, pady=10).pack(side="left")

        body = tk.Frame(dlg, bg=BG_MAIN)
        body.pack(fill="x", padx=18, pady=12)

        # Name field
        tk.Label(body, text="Expense Name", bg=BG_MAIN, fg=TEXT_DIM,
                 font=("Segoe UI", 8)).pack(anchor="w", pady=(0, 2))
        v_name = tk.StringVar(value=exp["name"])
        name_f = tk.Frame(body, bg=BG_CARD, highlightthickness=1,
                          highlightbackground=BORDER)
        name_f.pack(fill="x", pady=(0, 8))
        tk.Entry(name_f, textvariable=v_name, bg=BG_CARD, fg=TEXT_H,
                 insertbackground=TEXT_H, relief="flat",
                 font=("Segoe UI", 10), highlightthickness=0,
                 width=28).pack(fill="x", padx=8, pady=5)

        # Amount field
        tk.Label(body, text="Amount", bg=BG_MAIN, fg=TEXT_DIM,
                 font=("Segoe UI", 8)).pack(anchor="w", pady=(0, 2))
        v_amt = tk.StringVar(value=str(int(exp["amount"])))
        amt_f = tk.Frame(body, bg=BG_CARD, highlightthickness=1,
                         highlightbackground=BORDER)
        amt_f.pack(fill="x", pady=(0, 4))
        tk.Label(amt_f, text="₱", bg=BG_CARD, fg=ACCENT,
                 font=("Consolas", 10, "bold")).pack(side="left", padx=(8, 0))
        tk.Entry(amt_f, textvariable=v_amt, bg=BG_CARD, fg=TEXT_H,
                 insertbackground=TEXT_H, relief="flat",
                 font=("Consolas", 10), highlightthickness=0,
                 width=16).pack(side="left", padx=4, pady=5)

        # Buttons
        btn_f = tk.Frame(dlg, bg=BG_MAIN)
        btn_f.pack(fill="x", padx=18, pady=(0, 14))

        def _save():
            name = v_name.get().strip()
            if not name:
                messagebox.showwarning("Missing", "Enter an expense name.",
                                       parent=dlg)
                return
            try:
                amt = float(v_amt.get() or 0)
                if amt <= 0:
                    raise ValueError
            except ValueError:
                messagebox.showwarning("Invalid", "Enter a valid positive amount.",
                                       parent=dlg)
                return
            self._expenses[idx] = {"name": name, "amount": amt}
            self._persist()
            self._refresh_expenses()
            dlg.destroy()

        save_btn = tk.Button(btn_f, text="💾  Save", command=_save,
                             bg=BTN_GREEN, fg="white",
                             activebackground=BTN_GREEN_H,
                             relief="flat", font=("Segoe UI", 10, "bold"),
                             pady=8, cursor="hand2")
        save_btn.pack(fill="x", pady=(0, 6))
        hover_bind(save_btn, BTN_GREEN, BTN_GREEN_H)

        cancel_btn = tk.Button(btn_f, text="Cancel", command=dlg.destroy,
                               bg=BG_CARD, fg=TEXT_DIM,
                               activebackground=BG_PANEL,
                               relief="flat", font=("Segoe UI", 9),
                               pady=6, cursor="hand2",
                               highlightthickness=1, highlightbackground=BORDER)
        cancel_btn.pack(fill="x")

        dlg.bind("<Return>", lambda e: _save())
        dlg.bind("<Escape>", lambda e: dlg.destroy())

    def _delete_expense_at(self, idx):
        del self._expenses[idx]
        self._persist()
        self._refresh_expenses()

    def _add_expense(self):
        name = self._v_exp_name.get().strip()
        if not name:
            messagebox.showwarning("Missing", "Enter an expense name.", parent=self)
            return
        try:
            amt = float(self._v_exp_amount.get() or 0)
            if amt <= 0:
                raise ValueError
        except ValueError:
            messagebox.showwarning("Invalid", "Enter a valid positive amount.",
                                   parent=self)
            return
        self._expenses.append({"name": name, "amount": amt})
        self._v_exp_name.set("")
        self._v_exp_amount.set("")
        self._persist()
        self._refresh_expenses()

    def _selected_index(self):
        if self._active_tree is None:
            return None
        sel = self._active_tree.selection()
        return int(sel[0]) if sel else None

    def _edit_selected(self):
        idx = self._selected_index()
        if idx is None:
            messagebox.showinfo("Select", "Select a row to edit.")
            return
        rec = self._records[idx]

        # open session → load inline edit form
        if rec.get("session_type") == "open" and rec.get("time_out") == "OPEN":
            self._load_open_for_edit(idx, rec)
            return

        self._edit_idx    = idx
        self._editing_open = False

        # restore form fields
        pc_num = int(rec["pc"]) if str(rec["pc"]).isdigit() else 1
        self._select_form_pc(pc_num)
        self._v_name.set(rec["name"])
        self._v_amount.set(str(rec["amount"]))
        self._v_disc.set(str(rec.get("discount", 0)))
        self._v_comment.set(rec.get("comment", ""))

        def parse_time(ts, tp):
            try:
                parts = ts.split(); hm = parts[0].split(":")
                tp._hour.set(hm[0]); tp._min.set(hm[1])
                tp._ampm.set(parts[1] if len(parts) > 1 else "AM")
            except Exception: pass

        parse_time(rec["time_in"], self._tp_in)

        t_in  = parse_session_time(rec["time_in"])
        t_out = parse_session_time(rec["time_out"])
        if t_in and t_out:
            dur_min = int((t_out - t_in).total_seconds() // 60)
            dur_key = minutes_to_dur_input(dur_min)
            self._v_dur_input.set(dur_key)
            for v, btn in self._dur_btns.items():
                if v == "open":
                    btn.config(bg=BTN_ORANGE, fg="white")
                else:
                    btn.config(bg=ACCENT2 if v == dur_key else BG_CARD,
                               fg="white"  if v == dur_key else TEXT_MAIN)
        else:
            self._select_duration("1:00")

        self._amount_manual = True
        self._recalc()
        self._update_form_visibility()  # sets button to SAVE CHANGES

    def _load_open_for_edit(self, idx, rec):
        """Load an OPEN session record into the form for closing."""
        self._edit_idx     = idx
        self._editing_open = True

        pc_num = int(rec["pc"]) if str(rec["pc"]).isdigit() else 1
        self._select_form_pc(pc_num)
        self._v_name.set(rec["name"])
        self._v_comment.set(rec.get("comment", ""))

        def parse_time(ts, tp):
            try:
                parts = ts.split(); hm = parts[0].split(":")
                tp._hour.set(hm[0]); tp._min.set(hm[1])
                tp._ampm.set(parts[1] if len(parts) > 1 else "AM")
            except Exception: pass

        parse_time(rec["time_in"], self._tp_in)
        self._tp_out.set_now()   # default Time Out to current time

        self._amount_manual = False
        self._update_form_visibility()
        self._recalc_open_edit()

    # ── FORM HELPERS ────────────────────────────────────────────────────────
    def _clear_form(self):
        self._edit_idx     = None
        self._editing_open = False
        self._amount_manual = False
        self._select_form_pc(1)
        self._v_name.set("")
        self._tp_in.set_now()
        self._v_amount.set("0")
        self._v_disc.set("0")
        self._v_final.set("0")
        self._v_dur.set("—")
        self._v_timeout_display.set("—")
        self._v_comment.set("")
        self._select_duration("1:00")   # resets selected_dur, calls _update_form_visibility
        self.after(50, self._recalc)

    # ── BOOKINGS TAB ─────────────────────────────────────────────────────────
    def _build_bookings_tab(self, parent):
        parent.columnconfigure(0, weight=1)
        parent.rowconfigure(1, weight=1)

        # ── Booking Form ──────────────────────────────────────────────────────
        form = tk.Frame(parent, bg=BG_PANEL)
        form.grid(row=0, column=0, sticky="ew", pady=(0, 1))
        form.columnconfigure(1, weight=1)
        form.columnconfigure(3, weight=1)

        # Row 0: Name + Expected Time In
        tk.Label(form, text="Customer Name", bg=BG_PANEL, fg=TEXT_DIM,
                 font=("Segoe UI", 8)).grid(row=0, column=0, padx=(12, 4),
                                             pady=(10, 4), sticky="w")
        self._bk_v_name = tk.StringVar()
        nf = tk.Frame(form, bg=BG_CARD, highlightthickness=1,
                      highlightbackground=BORDER)
        nf.grid(row=0, column=1, sticky="ew", padx=(0, 12), pady=(10, 4))
        tk.Entry(nf, textvariable=self._bk_v_name, bg=BG_CARD, fg=TEXT_H,
                 insertbackground=TEXT_H, relief="flat",
                 font=("Segoe UI", 10), highlightthickness=0).pack(
            fill="x", padx=8, pady=4)

        tk.Label(form, text="Expected Time In", bg=BG_PANEL, fg=TEXT_DIM,
                 font=("Segoe UI", 8)).grid(row=0, column=2, padx=(0, 4),
                                             pady=(10, 4), sticky="w")
        self._bk_tp = TimePicker(form)
        self._bk_tp.set_now()
        self._bk_tp.grid(row=0, column=3, sticky="w", padx=(0, 12),
                         pady=(10, 4))

        # Row 1: Duration label
        tk.Label(form, text="DURATION", bg=BG_PANEL, fg=TEXT_DIM,
                 font=("Segoe UI", 7, "bold")).grid(
            row=1, column=0, columnspan=4, padx=12, pady=(4, 2), sticky="w")
        tk.Frame(form, bg=BORDER, height=1).grid(
            row=2, column=0, columnspan=4, sticky="ew", padx=12)

        # Row 3: Duration buttons
        dur_g = tk.Frame(form, bg=BG_PANEL)
        dur_g.grid(row=3, column=0, columnspan=4, sticky="ew", padx=12,
                   pady=(4, 4))
        for c in range(4):
            dur_g.columnconfigure(c, weight=1)

        self._bk_dur_btns    = {}
        self._bk_selected_dur = "1:00"
        for idx, (label, val) in enumerate(DURATION_PRESETS):
            btn = tk.Button(dur_g, text=label,
                            command=lambda v=val: self._bk_select_dur(v),
                            bg=ACCENT2 if val == "1:00" else BG_CARD,
                            fg="white"  if val == "1:00" else TEXT_MAIN,
                            relief="flat", font=("Segoe UI", 9, "bold"),
                            pady=4, cursor="hand2",
                            activebackground=ACCENT2, activeforeground="white")
            btn.grid(row=idx // 4, column=idx % 4, padx=2, pady=2, sticky="ew")
            self._bk_dur_btns[val] = btn
        open_btn = tk.Button(dur_g, text="∞  Open",
                             command=lambda: self._bk_select_dur("open"),
                             bg=BTN_ORANGE, fg="white", relief="flat",
                             font=("Segoe UI", 9, "bold"), pady=4, cursor="hand2",
                             activebackground=BTN_ORANGE_H, activeforeground="white")
        open_btn.grid(row=2, column=0, columnspan=4, padx=2, pady=2, sticky="ew")
        self._bk_dur_btns["open"] = open_btn

        # Row 4: PC selector label
        tk.Label(form, text="SELECT PC", bg=BG_PANEL, fg=TEXT_DIM,
                 font=("Segoe UI", 7, "bold")).grid(
            row=4, column=0, columnspan=4, padx=12, pady=(4, 2), sticky="w")
        tk.Frame(form, bg=BORDER, height=1).grid(
            row=5, column=0, columnspan=4, sticky="ew", padx=12)

        # Row 6: PC buttons
        pc_g = tk.Frame(form, bg=BG_PANEL)
        pc_g.grid(row=6, column=0, columnspan=4, sticky="ew", padx=12,
                  pady=(4, 4))
        for c in range(5):
            pc_g.columnconfigure(c, weight=1)

        self._bk_v_pc    = tk.StringVar(value="1")
        self._bk_pc_btns = {}
        for i in range(1, 16):
            btn = tk.Button(pc_g, text=str(i),
                            command=lambda n=i: self._bk_select_pc(n),
                            bg=ACCENT if i == 1 else BG_CARD,
                            fg="white"  if i == 1 else TEXT_MAIN,
                            relief="flat", font=("Segoe UI", 9, "bold"),
                            width=3, pady=3, cursor="hand2",
                            activebackground=ACCENT, activeforeground="white")
            btn.grid(row=(i-1)//5, column=(i-1)%5, padx=2, pady=2, sticky="ew")
            self._bk_pc_btns[i] = btn

        # Row 7: Deposit + Add Booking button
        r7 = tk.Frame(form, bg=BG_PANEL)
        r7.grid(row=7, column=0, columnspan=4, sticky="ew", padx=12,
                pady=(4, 10))
        r7.columnconfigure(1, weight=1)

        tk.Label(r7, text="Deposit (optional)", bg=BG_PANEL, fg=TEXT_DIM,
                 font=("Segoe UI", 8)).grid(row=0, column=0, padx=(0, 4),
                                             sticky="w")
        self._bk_v_deposit = tk.StringVar()
        dep_f = tk.Frame(r7, bg=BG_CARD, highlightthickness=1,
                         highlightbackground=BORDER)
        dep_f.grid(row=0, column=1, sticky="ew", padx=(0, 12))
        tk.Label(dep_f, text="₱", bg=BG_CARD, fg=ACCENT,
                 font=("Consolas", 10, "bold")).pack(side="left", padx=(6, 0))
        tk.Entry(dep_f, textvariable=self._bk_v_deposit, bg=BG_CARD, fg=TEXT_H,
                 insertbackground=TEXT_H, relief="flat",
                 font=("Consolas", 10), highlightthickness=0, width=10).pack(
            side="left", padx=4, pady=4)

        add_btn = tk.Button(r7, text="＋  Add Booking",
                            command=self._add_booking,
                            bg=ACCENT2, fg="white",
                            activebackground=ACCENT2_H,
                            relief="flat", font=("Segoe UI", 10, "bold"),
                            padx=14, pady=6, cursor="hand2")
        add_btn.grid(row=0, column=2)
        hover_bind(add_btn, ACCENT2, ACCENT2_H)

        # ── Bookings List ──────────────────────────────────────────────────────
        list_f = tk.Frame(parent, bg=BG_MAIN)
        list_f.grid(row=1, column=0, sticky="nsew")
        list_f.columnconfigure(0, weight=1)
        list_f.rowconfigure(1, weight=1)

        # Column header
        hdr = tk.Frame(list_f, bg=BG_CARD)
        hdr.grid(row=0, column=0, sticky="ew")
        _BK_COLS = [("PC", 35), ("Name", 130), ("Expected Time", 105),
                    ("Duration", 75), ("Deposit", 65), ("Status", 75),
                    ("Actions", 175)]
        for col_txt, col_w in _BK_COLS:
            tk.Label(hdr, text=col_txt, bg=BG_CARD, fg=TEXT_DIM,
                     font=("Segoe UI", 8, "bold"), width=col_w // 7,
                     anchor="center").pack(side="left", padx=4, pady=5)

        # Scrollable rows canvas
        scrl_f = tk.Frame(list_f, bg=BG_MAIN)
        scrl_f.grid(row=1, column=0, sticky="nsew")
        scrl_f.columnconfigure(0, weight=1)
        scrl_f.rowconfigure(0, weight=1)

        self._bk_canvas = tk.Canvas(scrl_f, bg=BG_MAIN, highlightthickness=0)
        vsb_bk = ttk.Scrollbar(scrl_f, orient="vertical",
                                command=self._bk_canvas.yview)
        self._bk_canvas.configure(yscrollcommand=vsb_bk.set)
        self._bk_canvas.grid(row=0, column=0, sticky="nsew")
        vsb_bk.grid(row=0, column=1, sticky="ns")

        self._bk_list_frame = tk.Frame(self._bk_canvas, bg=BG_MAIN)
        self._bk_canvas_win = self._bk_canvas.create_window(
            (0, 0), window=self._bk_list_frame, anchor="nw")
        self._bk_list_frame.bind(
            "<Configure>",
            lambda e: self._bk_canvas.configure(
                scrollregion=self._bk_canvas.bbox("all")))
        self._bk_canvas.bind(
            "<Configure>",
            lambda e: self._bk_canvas.itemconfig(
                self._bk_canvas_win, width=e.width))

    # Column pixel widths (must match header above)
    _BK_COL_W = [35, 130, 105, 75, 65, 75, 175]

    def _refresh_bookings(self):
        for w in self._bk_list_frame.winfo_children():
            w.destroy()
        pending  = [b for b in self._bookings if b.get("status") == "pending"]
        started  = [b for b in self._bookings if b.get("status") == "started"]
        shown    = pending + started  # cancelled are removed from list
        if not shown:
            tk.Label(self._bk_list_frame, text="No bookings yet.",
                     bg=BG_MAIN, fg=TEXT_DIM,
                     font=("Segoe UI", 9)).pack(pady=20)
            return
        for i, b in enumerate(shown):
            bg = ROW_ODD if i % 2 == 0 else ROW_EVEN
            row = tk.Frame(self._bk_list_frame, bg=bg)
            row.pack(fill="x")

            status = b.get("status", "pending")
            status_fg = "#e3b341" if status == "pending" else BTN_GREEN_H

            cells = [
                (str(b.get("pc", "")),    self._BK_COL_W[0], "center"),
                (b.get("name", ""),       self._BK_COL_W[1], "w"),
                (b.get("exp_time", ""),   self._BK_COL_W[2], "center"),
                (b.get("duration", ""),   self._BK_COL_W[3], "center"),
                (f"₱{b.get('deposit', 0):.0f}" if b.get("deposit", 0) > 0 else "—",
                 self._BK_COL_W[4], "center"),
            ]
            for txt, w, anchor in cells:
                tk.Label(row, text=txt, bg=bg, fg=TEXT_MAIN,
                         font=("Segoe UI", 9), width=w // 7,
                         anchor=anchor).pack(side="left", padx=4, pady=6)

            # Status label
            tk.Label(row, text=status.capitalize(), bg=bg, fg=status_fg,
                     font=("Segoe UI", 9, "bold"),
                     width=self._BK_COL_W[5] // 7,
                     anchor="center").pack(side="left", padx=4)

            # Action buttons
            act_f = tk.Frame(row, bg=bg)
            act_f.pack(side="left", padx=4)
            if status == "pending":
                start_btn = tk.Button(
                    act_f, text="▶ Start",
                    command=lambda bk=b: self._start_booking_session(bk),
                    bg=BTN_GREEN, fg="white",
                    activebackground=BTN_GREEN_H,
                    relief="flat", font=("Segoe UI", 8, "bold"),
                    padx=6, pady=2, cursor="hand2")
                start_btn.pack(side="left", padx=(0, 4))
                hover_bind(start_btn, BTN_GREEN, BTN_GREEN_H)

                cancel_btn = tk.Button(
                    act_f, text="✕ Cancel",
                    command=lambda bk=b: self._cancel_booking(bk),
                    bg=BG_CARD, fg=ACCENT,
                    activebackground=ACCENT,
                    activeforeground="white",
                    relief="flat", font=("Segoe UI", 8),
                    padx=6, pady=2, cursor="hand2",
                    highlightthickness=1, highlightbackground=BORDER)
                cancel_btn.pack(side="left")
                hover_bind(cancel_btn, BG_CARD, ACCENT)

    def _bk_select_dur(self, val):
        self._bk_selected_dur = val
        for v, btn in self._bk_dur_btns.items():
            if v == "open":
                btn.config(bg=BTN_ORANGE_H if v == val else BTN_ORANGE,
                           fg="white")
            else:
                btn.config(bg=ACCENT2 if v == val else BG_CARD,
                           fg="white"  if v == val else TEXT_MAIN)

    def _bk_select_pc(self, pc_num):
        self._bk_v_pc.set(str(pc_num))
        for n, btn in self._bk_pc_btns.items():
            if n == pc_num:
                btn.config(bg=ACCENT, fg="white")
            elif n in self._shutdown_pcs:
                btn.config(bg="#1a1a1a", fg="#404040")
            else:
                btn.config(bg=BG_CARD, fg=TEXT_MAIN)

    def _refresh_bk_pc_btns(self):
        """Update booking form PC button colours to reflect current shutdown state."""
        try:
            sel = int(self._bk_v_pc.get())
        except ValueError:
            sel = 1
        self._bk_select_pc(sel)

    def _add_booking(self):
        name = self._bk_v_name.get().strip()
        if not name:
            messagebox.showwarning("Missing", "Enter a customer name.", parent=self)
            return
        pc = self._bk_v_pc.get()
        if int(pc) in self._shutdown_pcs:
            messagebox.showwarning("PC Shutdown",
                                   f"PC {pc} is shutdown and cannot be booked.",
                                   parent=self)
            return
        try:
            deposit = float(self._bk_v_deposit.get() or 0)
            if deposit < 0:
                raise ValueError
        except ValueError:
            messagebox.showwarning("Invalid", "Enter a valid deposit amount (0 or more).",
                                   parent=self)
            return

        exp_time_str, exp_time_min = self._bk_tp.get_time()
        dur = self._bk_selected_dur
        dur_label = "Open" if dur == "open" else next(
            (lbl for lbl, v in DURATION_PRESETS if v == dur), dur)

        # Generate booking ID
        prefix = f"BK-{date.today().strftime('%Y%m%d')}-"
        max_seq = 0
        for bk in self._bookings:
            bid = bk.get("id", "")
            if bid.startswith(prefix):
                try:
                    max_seq = max(max_seq, int(bid[len(prefix):]))
                except ValueError:
                    pass
        booking = {
            "id":               f"{prefix}{max_seq + 1:03d}",
            "pc":               pc,
            "name":             name,
            "exp_time":         exp_time_str,
            "exp_time_minutes": exp_time_min,
            "duration":         dur_label,
            "duration_key":     dur,
            "duration_minutes": parse_dur_input(dur) if dur != "open" else 0,
            "deposit":          deposit,
            "status":           "pending",
        }
        self._bookings.append(booking)
        self._bk_v_name.set("")
        self._bk_v_deposit.set("")
        self._persist()
        self._refresh_bookings()
        self._update_pc_grid()

    def _start_booking_session(self, booking):
        """Pre-fill the main session form from a booking and switch to it."""
        pc_num = int(booking["pc"]) if str(booking["pc"]).isdigit() else 1
        self._select_form_pc(pc_num)
        self._v_name.set(booking["name"])

        # Set Expected Time In
        try:
            parts = booking["exp_time"].split()
            hm    = parts[0].split(":")
            self._tp_in._hour.set(hm[0])
            self._tp_in._min.set(hm[1])
            self._tp_in._ampm.set(parts[1] if len(parts) > 1 else "AM")
        except Exception:
            self._tp_in.set_now()

        # Set duration
        dur_key = booking.get("duration_key", "1:00")
        self._select_duration(dur_key)

        # Subtract deposit from auto-calculated amount
        deposit = booking.get("deposit", 0)
        if deposit > 0 and dur_key != "open":
            dur_min  = booking.get("duration_minutes", 0)
            rate_key = f"rate_{dur_key}"
            full_amt = (self._cfg[rate_key] if rate_key in self._cfg
                        else calc_amount(dur_min, self._cfg))
            net_amt  = max(0, full_amt - deposit)
            self._v_amount.set(str(int(net_amt)))
            self._amount_manual = True

        # Mark as started and clear its notification
        booking["status"] = "started"
        self._remove_notification(f"bk_{booking.get('id', '')}")
        self._persist()
        self._refresh_bookings()
        self._update_pc_grid()

        # Switch to main form
        self._switch_tab("current")

    def _cancel_booking(self, booking):
        self._bookings.remove(booking)
        self._persist()
        self._refresh_bookings()
        self._update_pc_grid()

    # ── AI ASSISTANT PANEL ───────────────────────────────────────────────────
    def _build_ai_panel(self):
        panel = tk.Frame(self, bg=BG_PANEL, width=340)
        panel.grid(row=0, column=1, rowspan=4, sticky="nsew")
        panel.grid_propagate(False)
        panel.columnconfigure(0, weight=1)
        panel.rowconfigure(1, weight=1)
        self._ai_panel_frame = panel
        panel.grid_remove()          # hidden by default

        # ── Panel header ──────────────────────────────────────────────────
        hdr = tk.Frame(panel, bg=BG_CARD)
        hdr.grid(row=0, column=0, sticky="ew")
        hdr.columnconfigure(0, weight=1)
        tk.Label(hdr, text="🤖  AI Assistant", bg=BG_CARD, fg=TEXT_H,
                 font=("Segoe UI", 11, "bold")).grid(
            row=0, column=0, padx=12, pady=8, sticky="w")
        tk.Button(hdr, text="✕", command=self._toggle_ai_panel,
                  bg=BG_CARD, fg=TEXT_DIM, relief="flat",
                  font=("Segoe UI", 10), cursor="hand2",
                  activebackground=ACCENT, activeforeground="white").grid(
            row=0, column=1, padx=8)
        tk.Frame(hdr, bg=ACCENT2, height=2).grid(
            row=1, column=0, columnspan=2, sticky="ew")

        # API key notice (shown only if key is empty)
        self._ai_notice_frame = tk.Frame(panel, bg="#1a1000")
        self._ai_notice_frame.grid(row=2, column=0, sticky="ew")
        self._ai_notice_lbl = tk.Label(
            self._ai_notice_frame,
            text="⚠  OpenRouter API key nahi hai.\nconfig.json mein 'openrouter_api_key' add karo.",
            bg="#1a1000", fg="#e3b341",
            font=("Segoe UI", 8), justify="left", wraplength=310)
        self._ai_notice_lbl.pack(padx=10, pady=6, anchor="w")

        # ── Messages area ─────────────────────────────────────────────────
        msg_area = tk.Frame(panel, bg=BG_MAIN)
        msg_area.grid(row=1, column=0, sticky="nsew")
        msg_area.columnconfigure(0, weight=1)
        msg_area.rowconfigure(0, weight=1)

        self._chat_canvas = tk.Canvas(msg_area, bg=BG_MAIN, highlightthickness=0)
        vsb = ttk.Scrollbar(msg_area, orient="vertical",
                            command=self._chat_canvas.yview)
        self._chat_canvas.configure(yscrollcommand=vsb.set)
        self._chat_canvas.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")

        self._chat_inner = tk.Frame(self._chat_canvas, bg=BG_MAIN)
        self._chat_win = self._chat_canvas.create_window(
            (0, 0), window=self._chat_inner, anchor="nw")
        self._chat_inner.bind(
            "<Configure>",
            lambda e: self._chat_canvas.configure(
                scrollregion=self._chat_canvas.bbox("all")))
        self._chat_canvas.bind(
            "<Configure>",
            lambda e: self._chat_canvas.itemconfig(
                self._chat_win, width=e.width))

        # ── Input area ────────────────────────────────────────────────────
        inp_outer = tk.Frame(panel, bg=BG_CARD)
        inp_outer.grid(row=3, column=0, sticky="ew")
        inp_outer.columnconfigure(0, weight=1)

        inp_border = tk.Frame(inp_outer, bg=BG_CARD, highlightthickness=1,
                              highlightbackground=BORDER)
        inp_border.grid(row=0, column=0, columnspan=2, sticky="ew",
                        padx=8, pady=8)
        inp_border.columnconfigure(0, weight=1)

        self._chat_input = tk.Text(
            inp_border, bg=BG_CARD, fg=TEXT_H,
            insertbackground=TEXT_H, relief="flat",
            font=("Segoe UI", 10), height=3, wrap="word",
            highlightthickness=0)
        self._chat_input.grid(row=0, column=0, sticky="ew", padx=6, pady=6)
        self._chat_input.bind("<Return>",       self._on_chat_enter)
        self._chat_input.bind("<Shift-Return>", lambda e: None)

        send_btn = tk.Button(inp_outer, text="Send ▸",
                             command=self._send_chat,
                             bg=ACCENT2, fg="white",
                             activebackground=ACCENT2_H,
                             relief="flat", font=("Segoe UI", 9, "bold"),
                             padx=12, pady=5, cursor="hand2")
        send_btn.grid(row=1, column=0, sticky="e", padx=8, pady=(0, 8))
        hover_bind(send_btn, ACCENT2, ACCENT2_H)

        self._update_ai_notice()

    def _update_ai_notice(self):
        key = self._cfg.get("openrouter_api_key", "").strip()
        if key:
            self._ai_notice_frame.grid_remove()
        else:
            self._ai_notice_frame.grid()

    def _toggle_ai_panel(self):
        if self._ai_panel_open:
            self._ai_panel_frame.grid_remove()
            self._ai_panel_open = False
            self._ai_toggle_btn.config(text="🤖  AI Assistant")
        else:
            self._update_ai_notice()
            self._ai_panel_frame.grid()
            self._ai_panel_open = True
            self._ai_toggle_btn.config(text="✕  Close AI")
            self._scroll_chat()

    # ── Chat UI helpers ───────────────────────────────────────────────────
    def _add_chat_message(self, role, text, timestamp=None):
        if timestamp is None:
            timestamp = datetime.now().strftime("%I:%M %p")
        self._chat_history.append(
            {"role": role, "content": text, "timestamp": timestamp})

        is_user  = (role == "user")
        bubble_bg = "#1d3557" if is_user else BG_CARD
        pack_side = "right" if is_user else "left"

        outer = tk.Frame(self._chat_inner, bg=BG_MAIN)
        outer.pack(fill="x", padx=6, pady=3)

        bubble = tk.Frame(outer, bg=bubble_bg, padx=10, pady=7)
        bubble.pack(side=pack_side, anchor="e" if is_user else "w",
                    padx=(60 if is_user else 0, 0 if is_user else 60))
        tk.Label(bubble, text=text, bg=bubble_bg, fg=TEXT_H,
                 font=("Segoe UI", 9), wraplength=220,
                 justify="left", anchor="w").pack(anchor="w")
        tk.Label(outer, text=timestamp, bg=BG_MAIN, fg=TEXT_DIM,
                 font=("Segoe UI", 7)).pack(
            side=pack_side, padx=4, pady=(0, 2), anchor="e" if is_user else "w")

        self._scroll_chat()
        if role == "assistant":
            self._check_ai_action(text)

    def _scroll_chat(self):
        self._chat_canvas.update_idletasks()
        self._chat_canvas.yview_moveto(1.0)

    def _show_typing(self):
        self._typing_frame = tk.Frame(self._chat_inner, bg=BG_MAIN)
        self._typing_frame.pack(fill="x", padx=6, pady=3, anchor="w")
        self._typing_lbl = tk.Label(
            self._typing_frame, text="●", bg=BG_CARD, fg=TEXT_DIM,
            font=("Segoe UI", 10), padx=10, pady=6)
        self._typing_lbl.pack(side="left")
        self._typing_dots = 0
        self._animate_typing()
        self._scroll_chat()

    def _animate_typing(self):
        if not self._ai_typing:
            return
        dots = ("●", "● ●", "● ● ●")[self._typing_dots % 3]
        if hasattr(self, "_typing_lbl") and self._typing_lbl.winfo_exists():
            self._typing_lbl.config(text=dots)
        self._typing_dots += 1
        self.after(380, self._animate_typing)

    def _hide_typing(self):
        self._ai_typing = False
        if hasattr(self, "_typing_frame") and self._typing_frame.winfo_exists():
            self._typing_frame.destroy()

    # ── Chat send / receive ───────────────────────────────────────────────
    def _on_chat_enter(self, event):
        if not (event.state & 0x1):   # Shift not held → send
            self._send_chat()
            return "break"

    def _send_chat(self):
        text = self._chat_input.get("1.0", "end-1c").strip()
        if not text:
            return
        self._chat_input.delete("1.0", "end")

        # Route through confirmation handler if pending
        if self._pending_record is not None:
            self._add_chat_message("user", text)
            self._handle_confirmation(text)
            return

        self._add_chat_message("user", text)

        # Client-side PC control commands (no API call needed)
        cmd = self._detect_pc_control_cmd(text)
        if cmd == "shutdown_available":
            self._exec_shutdown_available()
            return
        if cmd == "turnon_shutdown":
            self._exec_turnon_all_shutdown()
            return

        self._call_ai(text)

    def _detect_pc_control_cmd(self, text):
        """Return 'shutdown_available' or 'turnon_shutdown' if message matches, else None."""
        t = text.lower()
        has_pc   = "pc" in t
        has_off  = any(w in t for w in ("off", "band", "shutdown", "bند"))
        has_on   = any(w in t for w in ("on karo", "on kar", "chalu", "start", " on "))
        is_extra = any(w in t for w in ("extra", "available", "khali", "free",
                                         "use nahi", "jo use", "jo pc"))
        is_shutd = any(w in t for w in ("off", "band", "shutdown"))

        if has_pc and has_off and is_extra:
            return "shutdown_available"
        if has_pc and has_on and is_shutd:
            return "turnon_shutdown"
        # Also handle without "pc" keyword e.g. "extra band karo"
        if not has_pc:
            if is_extra and has_off:
                return "shutdown_available"
            if is_shutd and has_on:
                return "turnon_shutdown"
        return None

    def _exec_shutdown_available(self):
        """Shut down all currently available (free) PCs."""
        available = [i for i in range(1, 16)
                     if i not in self._shutdown_pcs
                     and self._get_active_session(i) is None]
        if not available:
            self._add_chat_message(
                "assistant",
                "Koi bhi PC available nahi tha — sab occupied ya already off hain.")
            return
        for pc in available:
            self._shutdown_pcs.add(pc)
        self._persist()
        self._update_pc_grid()
        pc_list = ", ".join(f"PC {n}" for n in available)
        self._add_chat_message(
            "assistant",
            f"{len(available)} PCs off kar diye: {pc_list}")

    def _exec_turnon_all_shutdown(self):
        """Turn on all shutdown PCs."""
        if not self._shutdown_pcs:
            self._add_chat_message("assistant", "Koi bhi PC off nahi tha.")
            return
        turned_on = sorted(self._shutdown_pcs)
        self._shutdown_pcs.clear()
        self._persist()
        self._update_pc_grid()
        pc_list = ", ".join(f"PC {n}" for n in turned_on)
        self._add_chat_message(
            "assistant",
            f"{len(turned_on)} PCs on kar diye: {pc_list}")

    def _build_context(self):
        now   = datetime.now()
        lines = [f"Current time: {now.strftime('%I:%M %p, %A %B %d %Y')}"]

        # Sessions
        lines.append(f"\nSessions today ({len(self._records)}):")
        for r in self._records:
            st = r.get("session_type", "fixed")
            lines.append(
                f"  PC {r['pc']} | {r['name']} | In:{r['time_in']} | "
                f"Out:{r.get('time_out','?')} | ₱{r.get('final',0):.0f} | {st}")

        closed = [r for r in self._records if r.get("session_type") != "open"]
        total  = sum(r.get("final", 0) for r in closed)
        lines.append(f"Total earnings: ₱{total:.0f}")

        # Expenses
        if self._expenses:
            lines.append(f"\nExpenses:")
            for e in self._expenses:
                lines.append(f"  {e['name']}: ₱{e['amount']:.0f}")
            lines.append(f"  Total: ₱{sum(e['amount'] for e in self._expenses):.0f}")

        # Edit log (last 15)
        if self._edit_log:
            lines.append(f"\nEdit log (last {min(15, len(self._edit_log))} changes):")
            for lg in self._edit_log[-15:]:
                lines.append(
                    f"  [{lg['timestamp']}] PC {lg['pc']} {lg['name']}: {lg['summary']}")

        # Pending bookings
        pending_bk = [b for b in self._bookings if b.get("status") == "pending"]
        if pending_bk:
            lines.append(f"\nPending bookings:")
            for b in pending_bk:
                lines.append(
                    f"  PC {b['pc']} | {b['name']} | {b['exp_time']} | {b['duration']}")

        # Open sessions with elapsed time
        open_s = [r for r in self._records
                  if r.get("session_type") == "open" and r.get("time_out") == "OPEN"]
        if open_s:
            lines.append(f"\nOpen (running) sessions:")
            for r in open_s:
                t_in = parse_session_time(r["time_in"])
                if t_in:
                    el = int((now - t_in).total_seconds() // 60)
                    elapsed = f"{el//60}h {el%60}m" if el >= 60 else f"{el}m"
                else:
                    elapsed = "?"
                lines.append(f"  PC {r['pc']} | {r['name']} | running {elapsed}")

        # PC status
        on_pcs  = [i for i in range(1, 16) if i not in self._shutdown_pcs]
        off_pcs = sorted(self._shutdown_pcs)
        lines.append(f"\nPCs ON: {on_pcs}")
        if off_pcs:
            lines.append(f"PCs OFF: {off_pcs}")

        return "\n".join(lines)

    def _call_ai(self, user_text):
        api_key = self._cfg.get("openrouter_api_key", "").strip()
        if not api_key:
            self._add_chat_message(
                "assistant",
                "OpenRouter API key config.json mein add karo.\n"
                "'openrouter_api_key' field mein apni key paste karo.")
            return

        self._ai_typing = True
        self._show_typing()

        context = self._build_context()
        system_prompt = (
            "You are a helpful cafe manager assistant.\n"
            "IMPORTANT LANGUAGE RULE:\n"
            "- NEVER write in Urdu script (never use arabic/urdu characters).\n"
            "- ALWAYS respond in Roman Urdu (Urdu words written in English alphabets).\n"
            "- Example of Roman Urdu: 'Aaj ka total 1500 rupay hai', "
            "'PC 3 pe ali 1 ghante se baitha hai'.\n"
            "- If user writes in English, reply in English.\n"
            "- If user writes in Roman Urdu, reply in Roman Urdu.\n"
            "- Never mix Urdu script characters in any response.\n\n"
            "IMPORTANT: When the user wants to ADD a session record "
            "(e.g. 'PC 3 pe ali ka 1 ghanta laga do' or 'add Ahmed to PC 5 for 2 hours'), "
            "extract PC number, name, and duration, calculate the amount from rates, "
            "then include this EXACT marker in your response (on its own line):\n"
            "[ADD_RECORD:{\"pc\":\"3\",\"name\":\"Ali\","
            "\"duration_key\":\"1:00\",\"amount\":90}]\n"
            "Then ask the user to confirm with yes/haan or no/nahi.\n\n"
            "SPECIAL COMMANDS (handled instantly by the app, no marker needed):\n"
            "- 'extra pc off karo' / 'available pc band karo' → turns off all free PCs\n"
            "- 'extra pc on karo' / 'band pc on karo' → turns on all shutdown PCs\n"
            "If user asks about these, tell them to just type the command.\n\n"
            "Available duration keys: 0:30, 1:00, 1:30, 2:00, 2:30, 3:00, 4:00, 5:00\n"
            f"Current rates: {json.dumps({k:v for k,v in self._cfg.items() if k.startswith('rate_')})}\n\n"
            "Current cafe data:\n" + context
        )

        # Build message list from history (last 20 messages)
        messages = [{"role": "system", "content": system_prompt}]
        for msg in self._chat_history[-20:]:
            if msg["role"] in ("user", "assistant"):
                messages.append({"role": msg["role"], "content": msg["content"]})

        # Models tried in order; skips 429/provider errors and tries next
        _FALLBACK_MODELS = [
            "nvidia/nemotron-nano-9b-v2:free",
            "meta-llama/llama-3.3-70b-instruct:free",
            "mistralai/mistral-small-3.1-24b-instruct:free",
            "nousresearch/hermes-3-llama-3.1-405b:free",
            "openrouter/auto",
        ]

        def _do_request():
            last_err = "Koi bhi model available nahi hai. Thodi der baad try karo."
            for model in _FALLBACK_MODELS:
                try:
                    payload = json.dumps({
                        "model":    model,
                        "messages": messages,
                    }).encode("utf-8")
                    req = urllib.request.Request(
                        "https://openrouter.ai/api/v1/chat/completions",
                        data=payload,
                        headers={
                            "Authorization": f"Bearer {api_key}",
                            "Content-Type":  "application/json",
                            "HTTP-Referer":  "cafe-manager-app",
                            "X-Title":       "Cafe Manager",
                        },
                        method="POST",
                    )
                    with urllib.request.urlopen(req, timeout=30) as resp:
                        data = json.loads(resp.read().decode("utf-8"))
                    reply = data["choices"][0]["message"]["content"].strip()
                    self.after(0, lambda r=reply: self._on_ai_response(r))
                    return                          # success — stop trying
                except urllib.error.HTTPError as e:
                    body = e.read().decode("utf-8", errors="ignore")
                    try:
                        last_err = json.loads(body)["error"]["message"]
                    except Exception:
                        last_err = body[:200]
                    # 429 / provider error → try next model; anything else → stop
                    if e.code not in (429, 400):
                        break
                except urllib.error.URLError:
                    last_err = "Internet connection nahi hai. Network check karo."
                    break
                except Exception as ex:
                    last_err = str(ex)
                    break
            self.after(0, lambda m=last_err: self._on_ai_error(m))

        threading.Thread(target=_do_request, daemon=True).start()

    def _on_ai_response(self, reply):
        self._hide_typing()
        self._add_chat_message("assistant", reply)

    def _on_ai_error(self, msg):
        self._hide_typing()
        self._add_chat_message("assistant", msg)

    def _check_ai_action(self, text):
        """Detect [ADD_RECORD:{...}] marker in AI response and store as pending."""
        match = re.search(r'\[ADD_RECORD:(\{[^}]+\})\]', text)
        if match:
            try:
                self._pending_record = json.loads(match.group(1))
            except Exception:
                self._pending_record = None

    def _handle_confirmation(self, text):
        """Process yes/no for a pending AI-suggested record."""
        lower = text.lower().strip()
        if lower in ("yes", "haan", "ha", "han", "y", "ok", "confirm"):
            rec_data = self._pending_record
            self._pending_record = None
            try:
                pc       = str(rec_data.get("pc", "1"))
                name     = str(rec_data.get("name", "Guest"))
                dur_key  = rec_data.get("duration_key", "1:00")
                amt      = float(rec_data.get("amount", 0))
                dur_min  = parse_dur_input(dur_key)
                now        = datetime.now()
                # Round current minute UP to next 5-minute mark
                raw_h24    = now.hour
                m_rounded  = math.ceil(now.minute / 5) * 5
                if m_rounded >= 60:
                    m_rounded = 0
                    raw_h24  += 1
                    if raw_h24 >= 24:
                        raw_h24 = 0
                h12        = raw_h24 % 12 or 12
                ampm       = "AM" if raw_h24 < 12 else "PM"
                time_in_s  = f"{h12:02d}:{m_rounded:02d} {ampm}"
                # Time Out = rounded Time In + duration (minutes end in 0 or 5)
                time_in_m  = raw_h24 * 60 + m_rounded
                time_out_s = calc_timeout_str(time_in_m, dur_min)
                new_rec = {
                    "pc":           pc,
                    "name":         name,
                    "time_in":      time_in_s,
                    "time_out":     time_out_s,
                    "duration":     fmt_duration(dur_min),
                    "amount":       amt,
                    "discount":     0,
                    "final":        amt,
                    "session_type": "fixed",
                    "advance":      0,
                    "comment":      "Added via AI",
                }
                new_rec["id"] = next_record_id(self._records)
                self._records.append(new_rec)
                self._persist()
                self._refresh_table()
                self._update_summary()
                self._update_pc_grid()
                self._add_chat_message(
                    "assistant",
                    f"✅ Record add ho gaya!\n"
                    f"PC {pc} — {name} — {fmt_duration(dur_min)} — ₱{amt:.0f}")
            except Exception as ex:
                self._add_chat_message("assistant", f"Record add karne mein error: {ex}")
        else:
            self._pending_record = None
            self._add_chat_message("assistant",
                                   "Theek hai, record add nahi kiya.")

    # ── SETTINGS / FOLDER ───────────────────────────────────────────────────
    def _open_settings(self):
        def _after_save():
            self._cfg = load_config()
            self._update_ai_notice()
        SettingsDialog(self, self._cfg, _after_save)

    def _open_folder(self):
        import subprocess, sys
        if sys.platform == "win32":   os.startfile(DATA_DIR)
        elif sys.platform == "darwin": subprocess.Popen(["open", DATA_DIR])
        else:                          subprocess.Popen(["xdg-open", DATA_DIR])


# ══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    app = CafeApp()
    app.mainloop()
